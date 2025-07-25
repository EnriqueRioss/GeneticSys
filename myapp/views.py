

import base64
import re
from django.core.files.base import ContentFile

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.db.models import Q, Count
from django.utils import timezone
from django.template.loader import render_to_string
from django.core.exceptions import PermissionDenied
from functools import wraps
from django.urls import reverse
import csv
from datetime import datetime, date, time
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.conf import settings
from django.utils import timezone
from django.contrib.contenttypes.models import ContentType
from datetime import datetime, time
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.template.loader import render_to_string
import json
# CAMBIO IMPORTANTE: Importamos 'modelformset_factory' directamente aquí


#Importaciones reportlab 
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django import forms



from .models import (
    Genetistas, Propositos, HistoriasClinicas, InformacionPadres, ExamenFisico,
    Parejas, AntecedentesPersonales, DesarrolloPsicomotor, PeriodoNeonatal,
    AntecedentesFamiliaresPreconcepcionales,
    EvaluacionGenetica, DiagnosticoPresuntivo, PlanEstudio, Project, Task,Parejas,Autorizaciones,ArchivoPlanEstudio
)
from .forms import (
    ExtendedUserCreationForm, HistoriasForm, PropositosForm, PadresPropositoForm,
    AntecedentesDesarrolloNeonatalForm, AntecedentesPreconcepcionalesForm,
    ExamenFisicoForm, ParejaPropositosForm, EvaluacionGeneticaForm,
    LoginForm, CreateNewTask, CreateNewProject, ReportSearchForm, AdminUserCreationForm, EvaluacionGeneticaForm, DiagnosticoFormSet, PlanEstudioFormSet,PasswordResetAdminForm,AdminUserEditForm,AutorizacionForm,PlanEstudioEditForm,ArchivarHistoriaForm,PatientSearchForm,EnfermedadActualForm,GenealogiaForm
)

from django.forms.models import model_to_dict

from django.views.decorators.cache import patch_cache_control



# ===== INICIO: LÓGICA DE FLUJO DE PASOS DINÁMICO =====

# 1. Definimos todos los nodos posibles con su nombre, vista y etiqueta.
ALL_STEPS = {
    'info_general': {'label': 'Información General', 'view_name': 'historia_crear_editar', 'icon': 'fas fa-file'},
    'datos_proposito': {'label': 'Datos del Propósito', 'view_name': 'paciente_crear', 'icon': 'fas fa-user'},
    'datos_pareja': {'label': 'Datos de la Pareja', 'view_name': 'pareja_crear', 'icon': 'fas fa-users'},
    'datos_padres': {'label': 'Datos de los Padres', 'view_name': 'padres_proposito_crear', 'icon': 'fas fa-users'},
    'enfermedad_actual': {'label': 'Enfermedad Actual', 'view_name': 'enfermedad_actual_crear', 'icon': 'fas fa-file-contract'},
    'antecedentes_personales': {'label': 'Antecedentes Personales', 'view_name': 'antecedentes_personales_crear', 'icon': 'fas fa-people-group'},
    'antecedentes_preconcepcionales': {'label': 'Antecedentes Preconcepcionales', 'view_name': 'antecedentes_preconcepcionales_crear', 'icon': 'fas fa-file'},
    'examen_fisico': {'label': 'Examen Físico', 'view_name': 'examen_fisico_crear_editar', 'icon': 'fas fa-stethoscope'},
    'evaluacion': {'label': 'Evaluación Genética', 'view_name': 'evaluacion_genetica_crear_editar', 'icon': 'fas fa-dna'},
    'genealogia': {'label': 'Genealogía', 'view_name': 'genealogia_crear_editar', 'icon': 'fas fa-sitemap'},
    'autorizacion': {'label': 'Autorización', 'view_name': 'autorizaciones_crear', 'icon': 'fas fa-signature'},
}

# 2. Definimos los flujos de trabajo basados en el motivo de la consulta.
WORKFLOWS = {
    'Proposito-Diagnóstico': [
        'info_general', 'datos_proposito', 'datos_padres', 'enfermedad_actual', 
        'antecedentes_personales', 'antecedentes_preconcepcionales', 
        'examen_fisico', 'evaluacion', 'genealogia', 'autorizacion'
    ],
    'Pareja-Asesoramiento Prenupcial': [
        'info_general', 'datos_pareja', 'enfermedad_actual', 
        'antecedentes_personales', 'antecedentes_preconcepcionales', 
        'examen_fisico', 'evaluacion', 'genealogia', 'autorizacion'
    ],
    # ... Añade aquí los otros motivos de consulta para parejas
    'Pareja-Preconcepcional': [
        'info_general', 'datos_pareja', 'enfermedad_actual', 
        'antecedentes_personales', 'antecedentes_preconcepcionales', 
        'examen_fisico', 'evaluacion', 'genealogia', 'autorizacion'
    ],
    'Pareja-Prenatal': [
        'info_general', 'datos_pareja', 'enfermedad_actual', 
        'antecedentes_personales', 'antecedentes_preconcepcionales', 
        'examen_fisico', 'evaluacion', 'genealogia', 'autorizacion'
    ],
}
# ===== FIN: LÓGICA DE FLUJO DE PASOS DINÁMICO =====



def get_edad_display(fecha_nacimiento):
    """Calcula y formatea la edad en años o meses."""
    if not fecha_nacimiento:
        return "N/A"
    today = timezone.now().date()
    age_years = today.year - fecha_nacimiento.year - ((today.month, today.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    if age_years >= 2:
        return f"{age_years} años"
    else:
        # Calcula la diferencia en meses
        months = (today.year - fecha_nacimiento.year) * 12 + today.month - fecha_nacimiento.month
        # Ajusta si el día actual es anterior al día de nacimiento
        if today.day < fecha_nacimiento.day:
            months -= 1
        # Previene meses negativos para recién nacidos
        months = max(0, months)
        return f"{months} meses"
    
    
def clear_editing_session(request):
    """
    Función auxiliar para limpiar la sesión de edición si se inició
    desde la lista de 'ver_historias'.
    """
    if request.session.pop('source_is_edit_list', False):
        if 'historia_en_progreso_id' in request.session:
            del request.session['historia_en_progreso_id']
            messages.info(request, "La edición de la historia clínica ha sido cancelada.")


def never_cache_on_get(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        response = view_func(request, *args, **kwargs)
        if request.method == 'GET':
            patch_cache_control(
                response, 
                no_cache=True, 
                no_store=True, 
                must_revalidate=True,
                max_age=0
            )
            response['Expires'] = '0'
            response['Pragma'] = 'no-cache'
        return response
    return _wrapped_view


# --- Role-Based Access Decorators ---
def role_required(allowed_roles):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            
            try:
                if not hasattr(request.user, 'genetistas'):
                    if request.user.is_superuser:
                        Genetistas.objects.get_or_create(user=request.user)
                        messages.info(request, "Perfil de Genetista creado para superusuario. Por favor, revise y ajuste el rol si es necesario en el panel de administración.")
                    else:
                        messages.error(request, "No tiene un perfil de aplicación configurado. Contacte al administrador.")
                        logout(request)
                        return redirect('login')
                
                user_gen_profile = request.user.genetistas
                user_role = user_gen_profile.rol
                
                if not user_role:
                    messages.error(request, "Su perfil de usuario no tiene un rol asignado. Contacte al administrador.")
                    if not request.user.is_superuser: logout(request)
                    return redirect('login')

            except Genetistas.DoesNotExist:
                messages.error(request, "Error crítico: Perfil de Genetista no encontrado y no se pudo crear.")
                if not request.user.is_superuser: logout(request)
                return redirect('login')

            if user_role not in allowed_roles:
                if request.user.is_superuser and user_role == 'ADM' and 'ADM' in allowed_roles:
                    pass
                elif request.user.is_superuser and user_role == 'ADM' and 'ADM' not in allowed_roles and 'GEN' in allowed_roles and 'LEC' in allowed_roles:
                    pass
                else:
                    role_names = [dict(Genetistas.ROL_CHOICES).get(r, r) for r in allowed_roles]
                    messages.error(request, f"Acceso denegado. Se requiere rol: {', '.join(role_names)}.")
                    return redirect('index')

            if user_role == 'LEC' and not user_gen_profile.associated_genetista:
                pass
            
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

admin_required = role_required(['ADM'])
genetista_required = role_required(['GEN', 'ADM'])
lector_required = role_required(['LEC', 'ADM'])
genetista_or_admin_required = role_required(['GEN', 'ADM'])
all_roles_required = role_required(['GEN', 'ADM', 'LEC'])


# --- Main Clinical Views ---

@login_required
@genetista_or_admin_required
@never_cache
def crear_editar_historia(request, historia_id=None):
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    instance = None
    editing = False
    
    if historia_id:
        instance = get_object_or_404(HistoriasClinicas, pk=historia_id)
        editing = True
        # --- LÓGICA AÑADIDA ---
        # Marcar que la edición se inició desde la lista
        request.session['source_is_edit_list'] = True
        # Asegurarse de que la sesión de progreso esté sincronizada
        request.session['historia_en_progreso_id'] = instance.pk
    elif 'historia_en_progreso_id' in request.session:
        try:
            instance = HistoriasClinicas.objects.get(pk=request.session['historia_en_progreso_id'])
            editing = True
        except HistoriasClinicas.DoesNotExist:
            del request.session['historia_en_progreso_id']
            
    if editing and instance:
        user_profile = request.user.genetistas
        if user_profile.rol == 'GEN' and instance.genetista != user_profile:
            raise PermissionDenied("No tiene permiso para editar esta historia clínica.")

    if request.method == 'POST':
        form = HistoriasForm(request.POST, instance=instance)
        
        if form.is_valid():
            try:
                historia = form.save(commit=False)
                
                if not editing:
                    genetista_profile = request.user.genetistas
                    if genetista_profile.rol in ['GEN', 'ADM']:
                        historia.genetista = genetista_profile
                
                historia.save()
                
                request.session.pop('form_data', None)
                
                # MODIFICADO: Lógica para "Guardar Borrador"
                if 'save_draft' in request.POST:
                    request.session.pop('historia_en_progreso_id', None)
                    # --- AÑADIDO ---
                    request.session.pop('source_is_edit_list', None) # Limpiar también al guardar borrador
                    # --- FIN ---
                    messages.success(request, f"Borrador de la Historia Clínica N° {historia.numero_historia} guardado exitosamente.")
                    redirect_url = reverse('ver_historias') # Redirigir a la lista
                    if is_ajax:
                        return JsonResponse({'success': True, 'redirect_url': redirect_url})
                    return redirect(redirect_url)

                # Lógica para "Siguiente"
                request.session['historia_en_progreso_id'] = historia.pk
                action_verb = "actualizada" if editing else "creada"
                messages.success(request, f"Historia Clínica N° {historia.numero_historia} {action_verb} exitosamente.")
                

                #  ====== INICIO DE LA LOGICA DEL FLUJO DE TRABAJO DINAMICO ========

                motivo = form.cleaned_data['motivo_tipo_consulta']

                # 1. Obtenemos el flujo de trabajo correspondiente y lo guardamos en la sesión
                workflow = WORKFLOWS.get(motivo, [])
                request.session['historia_workflow'] = workflow
                
                # 2. Obtenemos la URL de redirección para el SIGUIENTE paso del flujo
                # El siguiente paso es el segundo elemento de la lista (índice 1)
                next_node_name = workflow[1] if len(workflow) > 1 else None 

                if next_node_name:
                    # Buscamos el nombre de la vista en nuestro diccionario ALL_STEPS
                    next_view_name = ALL_STEPS[next_node_name]['view_name']
                    # Construimos los argumentos para reverse()
                    # Para 'paciente_crear' y 'pareja_crear', el único kwarg es 'historia_id'
                    next_kwargs = {'historia_id': historia.historia_id}
                    redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                else:
                    # Si no hay siguiente paso, redirigimos al índice (esto no debería pasar en un flujo normal)
                    redirect_url = reverse('index')


                # ===== FIN DE LA LÓGICA DEL FLUJO DE TRABAJO DINÁMICO =====

                
                # redirect_map = {
                # 'Proposito-Diagnóstico': ('paciente_crear', {'historia_id': historia.historia_id}),
                #   'Pareja-Asesoramiento Prenupcial': ('pareja_crear', {'historia_id': historia.historia_id}),
                #   'Pareja-Preconcepcional': ('pareja_crear', {'historia_id': historia.historia_id}),
                #   'Pareja-Prenatal': ('pareja_crear', {'historia_id': historia.historia_id}),
               # }
                #view_name, kwargs = redirect_map.get(motivo, ('index', {}))
                #redirect_url = reverse(view_name, kwargs=kwargs)


                if is_ajax:
                    return JsonResponse({'success': True, 'redirect_url': redirect_url})
                return redirect(redirect_url)

            except (IntegrityError, Genetistas.DoesNotExist) as e:
                error_msg = f"Error al guardar: {e}"
                if is_ajax: return JsonResponse({'success': False, 'errors': {'__all__': [error_msg]}}, status=400)
                messages.error(request, error_msg)
                request.session['form_data'] = request.POST.copy()
                return redirect(request.path_info)
        else:
            if is_ajax: return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            messages.error(request, "No se pudo guardar la historia. Corrija los errores.")
            request.session['form_data'] = request.POST.copy()
            return redirect(request.path_info)
            
    else:
        form = HistoriasForm(request.session.pop('form_data', None) or None, instance=instance)

    # ===== INICIO DE LA MODIFICACIÓN DEL CONTEXTO GET =====
    # Si estamos en modo de edición, intentamos recuperar el workflow de la sesión.
    # Si no, lo establecemos como una lista que solo contiene el paso actual.
    workflow = request.session.get('historia_workflow')
    if not workflow:
         # Creamos un workflow temporal para que la plantilla no falle al renderizar el stepper
        if instance and instance.motivo_tipo_consulta:
            workflow = WORKFLOWS.get(instance.motivo_tipo_consulta, ['info_general'])
        else:
            workflow = ['info_general']
        request.session['historia_workflow'] = workflow

    try:
        current_step_index = workflow.index('info_general') + 1
    except (ValueError, TypeError):
        current_step_index = 1

    context = {
        'form1': form, 
        'editing': editing, 
        'historia': instance,
        'current_node': 'info_general',            # <-- Le decimos a la plantilla cuál es el nodo actual
        'current_step_index': current_step_index,   # <-- Pasamos el índice numérico
        'ALL_STEPS': ALL_STEPS                      # <-- Pasamos el diccionario completo de pasos
    }

    # ===== FIN DE LA MODIFICACIÓN DEL CONTEXTO GET =====
    return render(request, "historia_clinica.html", context)

@login_required
@genetista_or_admin_required
@never_cache
def crear_paciente(request, historia_id):
    historia = get_object_or_404(HistoriasClinicas, historia_id=historia_id)
    user_profile = request.user.genetistas
    if user_profile.rol == 'GEN' and historia.genetista != user_profile:
        raise PermissionDenied("No tiene permiso para modificar pacientes de esta historia clínica.")
    
    existing_proposito = Propositos.objects.filter(historia=historia).first()
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    
    # --- AÑADIDO: Detectar si se viene de la gestión de pacientes ---
    from_gestion = request.GET.get('from_gestion') == 'true'


    # ===== INICIO DE LA LÓGICA DE CONTEXTO DINÁMICO =====
    current_node_name = 'datos_proposito'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 2 # Valor por defecto razonable

    context = {
        'form': None, 
        'historia': historia, 
        'editing': bool(existing_proposito), 
        'from_gestion': from_gestion,
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }
    # ===== FIN DE LA LÓGICA DE CONTEXTO DINÁMICO =====


    if request.method == 'POST':
        form = PropositosForm(request.POST, request.FILES, instance=existing_proposito)
        if form.is_valid():
            try:
                proposito = form.save(historia=historia)
                request.session.pop('form_data', None)

                # --- LÓGICA DE REDIRECCIÓN MODIFICADA ---
                if from_gestion:
                    messages.success(request, f"Paciente {proposito.nombres} {proposito.apellidos} actualizado exitosamente.")
                    redirect_url = reverse('gestion_pacientes')
                elif 'save_draft' in request.POST:
                    request.session.pop('historia_en_progreso_id', None)
                    messages.success(request, f"Borrador del paciente {proposito.nombres} guardado exitosamente.")
                    redirect_url = reverse('ver_historias') # Redirige a ver historias en lugar de index
                else:
                    action_verb = 'actualizado' if existing_proposito else 'creado'
                    messages.success(request, f"Paciente {proposito.nombres} {proposito.apellidos} {action_verb} exitosamente.")

                    # --- Lógica de redirección dinámica ---
                    next_step_index = current_step_index
                    if next_step_index < len(workflow):
                        next_node_name = workflow[next_step_index]
                        next_view_name = ALL_STEPS[next_node_name]['view_name']
                        # El siguiente paso ('datos_padres') necesita proposito_id
                        next_kwargs = {'historia_id': historia.historia_id, 'proposito_id': proposito.proposito_id}
                        redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                    else:
                        redirect_url = reverse('index')

                if is_ajax:
                    return JsonResponse({'success': True, 'redirect_url': redirect_url})
                return redirect(redirect_url)

            except Exception as e:
                error_msg = f"Error al guardar el paciente: {e}"
                if is_ajax: return JsonResponse({'success': False, 'errors': {'__all__': [error_msg]}}, status=400)
                messages.error(request, error_msg)
                request.session['form_data'] = request.POST.copy()
                return redirect(request.path_info)
        else:
            if is_ajax: return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            messages.error(request, "No se pudo guardar el paciente. Corrija los errores.")
            context['form'] = form # Añadimos el formulario con errores al contexto
            return render(request, "Crear_paciente.html", context)
    else: 
        form_data = request.session.pop('form_data', None)
        form = PropositosForm(form_data or None, instance=existing_proposito)
        if existing_proposito and not form_data:
            messages.info(request, f"Editando información para: {existing_proposito.nombres} {existing_proposito.apellidos}")
            
    # --- AÑADIDO: Pasar la variable al contexto ---
    context['form'] = form
    return render(request, "Crear_paciente.html", context)

# En tu archivo views.py
# In myapp/views.py

@login_required
@all_roles_required
def get_historia_details_ajax(request, historia_id):
    """
    Vista AJAX para obtener todos los detalles de una historia clínica para el modal.
    VERSIÓN MEJORADA: Maneja correctamente el caso de historias sin paciente.
    """
    historia = get_object_or_404(HistoriasClinicas.objects.select_related('genetista__user'), pk=historia_id)

    # ... (código de verificación de permisos y función format_value sin cambios) ...
    try:
        user_gen_profile = request.user.genetistas
        if user_gen_profile.rol == 'GEN' and historia.genetista != user_gen_profile:
            raise PermissionDenied("No tiene permiso para ver esta historia.")
        if user_gen_profile.rol == 'LEC':
            if not user_gen_profile.associated_genetista or historia.genetista != user_gen_profile.associated_genetista:
                raise PermissionDenied("No tiene permiso para ver esta historia.")
    except Genetistas.DoesNotExist:
        return JsonResponse({'error': 'Perfil de usuario no encontrado.'}, status=403)

    def format_value(value, default='N/A'):
        if value is None or value == '':
            return default
        if isinstance(value, datetime):
            return value.strftime('%d/%m/%Y %H:%M')
        if isinstance(value, date):
            return value.strftime('%d/%m/%Y')
        return str(value)

    propositos_qs = Propositos.objects.filter(historia=historia)
    
    if propositos_qs.exists():
        paciente_display_principal = ' y '.join([f"{p.nombres} {p.apellidos}" for p in propositos_qs])
    else:
        paciente_display_principal = "Sin Paciente Asignado"

    data = {
        'numero_historia': format_value(historia.numero_historia),
        'paciente_display_principal': paciente_display_principal,
        'motivo_consulta': format_value(historia.get_motivo_tipo_consulta_display()),
        'genetista_asignado': format_value(historia.genetista.user.get_full_name() if historia.genetista and historia.genetista.user else None),
        'medico_referencia': format_value(historia.medico),
        'centro_referencia': format_value(historia.centro_referencia),
        'fecha_creacion': format_value(historia.fecha_ingreso),
        'fecha_ultima_modificacion': format_value(historia.fecha_ultima_modificacion),
        'estado_display': historia.get_estado_display(),
        'estado_slug': historia.estado,
        'motivo_archivado': historia.motivo_archivado,
        'propositos': [],
        'is_pareja': propositos_qs.count() > 1,
    }
    
    # ===== INICIO DEL CÓDIGO A AÑADIR =====
    # Añadimos la URL de la genealogía a la respuesta JSON
    
    # 1. Preparamos el filtro para buscar los antecedentes familiares
    q_filter_antecedentes = Q()
    if data['is_pareja']:
        # Si es una pareja, buscamos el objeto Pareja
        pareja = Parejas.objects.filter(proposito_id_1__in=propositos_qs, proposito_id_2__in=propositos_qs).first()
        if pareja:
            q_filter_antecedentes = Q(pareja=pareja)
    elif propositos_qs.exists():
        # Si es un solo propósito, lo usamos para el filtro
        q_filter_antecedentes = Q(proposito=propositos_qs.first())

    # 2. Buscamos el registro de antecedentes y obtenemos la URL de la foto
    data['genealogia_url'] = None # Inicializamos por si no se encuentra
    if q_filter_antecedentes: # Solo si pudimos construir un filtro válido
        antecedente_familiar = AntecedentesFamiliaresPreconcepcionales.objects.filter(q_filter_antecedentes).first()
        if antecedente_familiar and antecedente_familiar.genealogia_foto:
            data['genealogia_url'] = antecedente_familiar.genealogia_foto.url
    # ===== FIN DEL CÓDIGO A AÑADIR =====

    # ... (el resto de la vista, incluyendo el bucle for p in propositos_qs, no necesita cambios) ...
    for p in propositos_qs:
        proposito_data = {
            'nombre_completo': f"{p.nombres} {p.apellidos}",
            'identificacion': format_value(p.identificacion),
            'lugar_nacimiento': format_value(p.lugar_nacimiento),
            'escolaridad': format_value(p.escolaridad),
            'ocupacion': format_value(p.ocupacion),
            'grupo_sanguineo': f"{format_value(p.grupo_sanguineo)} {format_value(p.factor_rh)}",
            'fecha_nacimiento': format_value(p.fecha_nacimiento),
            'padres_info': None,
            'examen_fisico': None,
        }

        if not data['is_pareja']:
            padres = InformacionPadres.objects.filter(proposito=p)
            padres_data = {}
            for padre_obj in padres:
                padres_data[padre_obj.tipo.lower()] = {
                    'nombre_completo': f"{padre_obj.nombres} {padre_obj.apellidos}",
                    'identificacion': format_value(padre_obj.identificacion),
                    'ocupacion': format_value(padre_obj.ocupacion),
                    'grupo_sanguineo': f"{format_value(padre_obj.grupo_sanguineo)} {format_value(padre_obj.factor_rh)}",
                }
            proposito_data['padres_info'] = padres_data

        ef = ExamenFisico.objects.filter(proposito=p).first()
        if ef:
            proposito_data['examen_fisico'] = {
                'talla': format_value(ef.talla),
                'peso': format_value(ef.peso),
                'medida_abrazada': format_value(ef.medida_abrazada),
                'segmento_superior': format_value(ef.segmento_superior),
                'segmento_inferior': format_value(ef.segmento_inferior),
                'distancia_interc_interna': format_value(ef.distancia_interc_interna),
                'distancia_interc_externa': format_value(ef.distancia_interc_externa),
                'distancia_interpupilar': format_value(ef.distancia_interpupilar),
            }

        data['propositos'].append(proposito_data)
        
    return JsonResponse(data)

# ... (todos los demás imports y vistas se mantienen igual) ...
@login_required
@all_roles_required
def get_historia_clinica_data(request, proposito_id):
    """
    Vista AJAX para obtener los datos de la historia clínica de un propósito específico.
    """
    proposito = get_object_or_404(_get_pacientes_queryset_for_role(request.user), pk=proposito_id)

    # El filtro Q se usará para varios modelos, así que lo definimos una vez
    q_filter = Q(proposito=proposito) | Q(pareja__proposito_id_1=proposito) | Q(pareja__proposito_id_2=proposito)

    evaluacion = EvaluacionGenetica.objects.filter(q_filter).prefetch_related('diagnosticos_presuntivos').first()
    antecedentes_personales = AntecedentesPersonales.objects.filter(q_filter).first()

    data = {
        'planes_de_estudio': [],
        'diagnosticos_presuntivos': [],
        'diagnostico_final': None,
        'evaluacion_id': None,
        # ===== CAMPOS NUEVOS AÑADIDOS =====
        'motivo_consulta': proposito.historia.get_motivo_tipo_consulta_display() if proposito.historia else 'N/A',
        'enfermedad_actual': antecedentes_personales.enfermedad_actual if antecedentes_personales and antecedentes_personales.enfermedad_actual else None,
        # ==================================
    }

    if evaluacion:
        data['evaluacion_id'] = evaluacion.evaluacion_id
        data['diagnosticos_presuntivos'] = list(evaluacion.diagnosticos_presuntivos.values_list('descripcion', flat=True))
        data['diagnostico_final'] = evaluacion.diagnostico_final
        
        planes = PlanEstudio.objects.filter(evaluacion=evaluacion).prefetch_related('archivos').order_by('fecha_visita', 'plan_id')
        
        for i, plan in enumerate(planes):
            archivos_list = [{
                'id': archivo.archivo_id,
                'nombre': archivo.get_display_name(),
                'url': archivo.archivo.url
            } for archivo in plan.archivos.all()]

            data['planes_de_estudio'].append({
                'id': plan.plan_id,
                'tipo_consulta': 'Consulta Inicial' if i == 0 else 'Consulta de Seguimiento',
                'completado': plan.completado,
                'accion': plan.accion,
                'asesoramiento_evoluciones': plan.asesoramiento_evoluciones or "",
                'fecha_visita': plan.fecha_visita.strftime('%Y-%m-%d') if plan.fecha_visita else None,
                'genetista_nombre': proposito.historia.genetista.user.get_full_name() if proposito.historia.genetista else "N/A",
                'archivos': archivos_list,
            })
    
    return JsonResponse(data)


@require_POST
@login_required
@genetista_or_admin_required
def edit_plan_estudio(request, plan_id):
    """
    Vista AJAX para editar un Plan de Estudio.
    """
    plan_instance = get_object_or_404(PlanEstudio, pk=plan_id)
    
    # Verificación de permisos
    user_profile = request.user.genetistas
    if user_profile.rol == 'GEN':
        genetista_historia = None
        if plan_instance.evaluacion.proposito:
            genetista_historia = plan_instance.evaluacion.proposito.historia.genetista
        elif plan_instance.evaluacion.pareja:
            genetista_historia = plan_instance.evaluacion.pareja.proposito_id_1.historia.genetista
        
        if genetista_historia != user_profile:
            return JsonResponse({'success': False, 'errors': {'__all__': ['No tiene permiso para editar este plan.']}}, status=403)

    form = PlanEstudioEditForm(request.POST, instance=plan_instance)

    if form.is_valid():
        plan_estudio = form.save(commit=False) # No guarda todavía, solo actualiza la instancia
        
        # Guardar la instancia principal
        plan_estudio.save()

        # Eliminar archivos marcados (ya manejado en el form.save)
        if form.cleaned_data.get('archivos_a_eliminar'):
            form.cleaned_data['archivos_a_eliminar'].delete()

        # Añadir nuevos archivos
        new_files = request.FILES.getlist('archivos_nuevos')
        for f in new_files:
            # Puedes añadir lógica aquí para el `nombre_descriptivo` si lo deseas
            ArchivoPlanEstudio.objects.create(plan_estudio=plan_estudio, archivo=f)

        return JsonResponse({'success': True, 'message': 'Plan de estudio actualizado correctamente.'})
    else:
        return JsonResponse({'success': False, 'errors': form.errors.as_json()}, status=400)



# Reemplaza esta función completa en views.py

@login_required
@genetista_or_admin_required
@never_cache
def crear_pareja(request, historia_id):
    historia = get_object_or_404(HistoriasClinicas, historia_id=historia_id)
    user_profile = request.user.genetistas
    if user_profile.rol == 'GEN' and historia.genetista != user_profile:
        raise PermissionDenied("No tiene permiso para modificar parejas de esta historia clínica.")

    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    from_gestion = request.GET.get('from_gestion') == 'true'

    propositos_en_historia = Propositos.objects.filter(historia=historia)
    pareja_existente = Parejas.objects.filter(
        proposito_id_1__in=propositos_en_historia,
        proposito_id_2__in=propositos_en_historia
    ).select_related('proposito_id_1', 'proposito_id_2').first()

    editing = bool(pareja_existente)


    # --- INICIO: LÓGICA DE CONTEXTO PARA EL STEPPER ---
    current_node_name = 'datos_pareja'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 2 # Fallback
    
    context = {
        'form': None, 
        'historia': historia, 
        'from_gestion': from_gestion,
        'editing': editing,
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }
    # --- FIN: LÓGICA DE CONTEXTO ---



    if request.method == 'POST':
        # La lógica POST no necesita cambios, ya que depende del form.is_valid()
        # y el form.clean() que ya estaban bien.
        form = ParejaPropositosForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    def get_or_create_proposito_from_form(form_cleaned_data, prefix_num, historia_obj, files_data):
                        identificacion = form_cleaned_data[f'identificacion_{prefix_num}']
                        
                        # Corrección: El modelo ya no tiene edad editable, se calcula en save()
                        proposito_data = {
                            'nombres': form_cleaned_data[f'nombres_{prefix_num}'],
                            'apellidos': form_cleaned_data[f'apellidos_{prefix_num}'],
                            'sexo': form_cleaned_data.get(f'sexo_{prefix_num}'),
                            'lugar_nacimiento': form_cleaned_data.get(f'lugar_nacimiento_{prefix_num}'),
                            'fecha_nacimiento': form_cleaned_data.get(f'fecha_nacimiento_{prefix_num}'),
                            'escolaridad': form_cleaned_data.get(f'escolaridad_{prefix_num}'),
                            'ocupacion': form_cleaned_data.get(f'ocupacion_{prefix_num}'),
                            'direccion': form_cleaned_data.get(f'direccion_{prefix_num}'),
                            'telefono': form_cleaned_data.get(f'telefono_{prefix_num}'),
                            'email': form_cleaned_data.get(f'email_{prefix_num}'),
                            'grupo_sanguineo': form_cleaned_data.get(f'grupo_sanguineo_{prefix_num}'),
                            'factor_rh': form_cleaned_data.get(f'factor_rh_{prefix_num}'),
                        }
                        
                        proposito_defaults = {k: v for k, v in proposito_data.items() if v is not None}
                        proposito_defaults['historia'] = historia_obj
                        
                        proposito, created = Propositos.objects.update_or_create(identificacion=identificacion, defaults=proposito_defaults)
                        
                        foto_file = files_data.get(f'foto_{prefix_num}')
                        if foto_file:
                            proposito.foto = foto_file
                        
                        # Si se usa update_or_create, el save del modelo no se llama si no hay cambios.
                        # Forzamos un save para asegurar que la edad se recalcule si la fecha de nacimiento cambió.
                        proposito.save() 
                        return proposito

                    proposito1 = get_or_create_proposito_from_form(form.cleaned_data, '1', historia, request.FILES)
                    proposito2 = get_or_create_proposito_from_form(form.cleaned_data, '2', historia, request.FILES)

                    if proposito1.pk == proposito2.pk:
                        raise IntegrityError("Los dos miembros de la pareja no pueden ser la misma persona.")

                    p_min, p_max = sorted([proposito1, proposito2], key=lambda p: p.pk)
                    
                    if pareja_existente:
                        pareja_existente.proposito_id_1 = p_min
                        pareja_existente.proposito_id_2 = p_max
                        pareja_existente.save()
                        pareja = pareja_existente
                        pareja_created = False
                    else:
                        pareja, pareja_created = Parejas.objects.get_or_create(
                            proposito_id_1=p_min,
                            proposito_id_2=p_max
                        )

                    ids_correctos = {proposito1.pk, proposito2.pk}
                    Propositos.objects.filter(historia=historia).exclude(pk__in=ids_correctos).delete()

                request.session.pop('form_data', None)


                # --- INICIO: LÓGICA DE REDIRECCIÓN DINÁMICA ---
                if from_gestion:
                    messages.success(request, f"Datos de la pareja actualizados exitosamente.")
                    redirect_url = reverse('gestion_pacientes')
                elif 'save_draft' in request.POST:
                    request.session.pop('historia_en_progreso_id', None)
                    request.session.pop('historia_workflow', None)
                    messages.success(request, f"Borrador de la pareja guardado exitosamente.")
                    redirect_url = reverse('ver_historias')
                else:
                    action_verb = 'creada' if pareja_created else 'actualizada'
                    messages.success(request, f"Pareja {action_verb} exitosamente.")
                    
                    next_node_name = workflow[current_step_index] if current_step_index < len(workflow) else None
                    if next_node_name:
                        next_view_name = ALL_STEPS[next_node_name]['view_name']
                        # El siguiente paso (ej. 'enfermedad_actual') necesita el ID de la pareja
                        next_kwargs = {'historia_id': historia.historia_id, 'tipo': 'pareja', 'objeto_id': pareja.pareja_id}
                        redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                    else:

                        redirect_url = reverse('index')
                # --- FIN: LÓGICA DE REDIRECCIÓN DINÁMICA ---

                if is_ajax:
                    return JsonResponse({'success': True, 'redirect_url': redirect_url})
                return redirect(redirect_url)
            except (IntegrityError, Exception) as e:
                error_msg = f"Error al guardar la pareja: {e}"
                if is_ajax: return JsonResponse({'success': False, 'errors': {'__all__': [error_msg]}}, status=400)
                messages.error(request, error_msg)
                context['form'] = form
                return render(request, 'Crear_pareja.html', context)
        else:
            if is_ajax: return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            messages.error(request, "No se pudo guardar la pareja. Corrija los errores.")
            context['form'] = form
            return render(request, 'Crear_pareja.html', context)
            
    else:  # Lógica GET
        form_data = request.session.pop('form_data', None)
        form_kwargs = {}

        if editing and not form_data:
            form_kwargs['conyuge1_instance'] = pareja_existente.proposito_id_1
            form_kwargs['conyuge2_instance'] = pareja_existente.proposito_id_2
            messages.info(request, f"Editando información para la pareja: {pareja_existente.proposito_id_1.nombres} y {pareja_existente.proposito_id_2.nombres}.")
        
        if form_data:
            form = ParejaPropositosForm(form_data)
        else:
            form = ParejaPropositosForm(**form_kwargs)
        
    context['form'] = form
    return render(request, 'Crear_pareja.html', context)


@login_required
@genetista_or_admin_required
@never_cache
def padres_proposito(request, historia_id, proposito_id):
    historia = get_object_or_404(HistoriasClinicas, historia_id=historia_id)
    proposito = get_object_or_404(Propositos, proposito_id=proposito_id, historia=historia)
    user_profile = request.user.genetistas
    if user_profile.rol == 'GEN' and proposito.historia.genetista != user_profile:
        raise PermissionDenied("No tiene permiso para modificar esta información.")

    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    

    # ===== INICIO DE LA LÓGICA DE CONTEXTO DINÁMICO =====
    current_node_name = 'datos_padres'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 3 # Valor por defecto

    context = {
        'form': None,
        'historia': historia,
        'proposito': proposito,
        'editing': False,
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }
    # ===== FIN DE LA LÓGICA DE CONTEXTO DINÁMICO =====



    padre_instance = InformacionPadres.objects.filter(proposito=proposito, tipo='Padre').first()
    madre_instance = InformacionPadres.objects.filter(proposito=proposito, tipo='Madre').first()
    editing = bool(padre_instance or madre_instance)

    if request.method == 'POST':
        form = PadresPropositoForm(request.POST, padre_instance=padre_instance, madre_instance=madre_instance)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # --- INICIO DE LA CORRECCIÓN --- ### <<<
                    # Construimos el diccionario para el padre explícitamente con los nombres de campo del MODELO
                    padre_defaults = {
                        'nombres': form.cleaned_data.get('padre_nombres'),
                        'apellidos': form.cleaned_data.get('padre_apellidos'),
                        'escolaridad': form.cleaned_data.get('padre_escolaridad'),
                        'ocupacion': form.cleaned_data.get('padre_ocupacion'),
                        'lugar_nacimiento': form.cleaned_data.get('padre_lugar_nacimiento'),
                        'fecha_nacimiento': form.cleaned_data.get('padre_fecha_nacimiento'),
                        'identificacion': form.cleaned_data.get('padre_identificacion'),
                        'grupo_sanguineo': form.cleaned_data.get('padre_grupo_sanguineo'),
                        'factor_rh': form.cleaned_data.get('padre_factor_rh'),
                        'telefono': form.cleaned_data.get('padre_telefono'),
                        'email': form.cleaned_data.get('padre_email'),
                        'direccion': form.cleaned_data.get('padre_direccion'),
                    }

                    # Construimos el diccionario para la madre explícitamente
                    madre_defaults = {
                        'nombres': form.cleaned_data.get('madre_nombres'),
                        'apellidos': form.cleaned_data.get('madre_apellidos'),
                        'escolaridad': form.cleaned_data.get('madre_escolaridad'),
                        'ocupacion': form.cleaned_data.get('madre_ocupacion'),
                        'lugar_nacimiento': form.cleaned_data.get('madre_lugar_nacimiento'),
                        'fecha_nacimiento': form.cleaned_data.get('madre_fecha_nacimiento'),
                        'identificacion': form.cleaned_data.get('madre_identificacion'),
                        'grupo_sanguineo': form.cleaned_data.get('madre_grupo_sanguineo'),
                        'factor_rh': form.cleaned_data.get('madre_factor_rh'),
                        'telefono': form.cleaned_data.get('madre_telefono'),
                        'email': form.cleaned_data.get('madre_email'),
                        'direccion': form.cleaned_data.get('madre_direccion'),
                    }

                    # Usamos update_or_create con los diccionarios corregidos
                    InformacionPadres.objects.update_or_create(
                        proposito=proposito, tipo='Padre', 
                        defaults={k: v for k, v in padre_defaults.items() if v is not None}
                    )
                    InformacionPadres.objects.update_or_create(
                        proposito=proposito, tipo='Madre', 
                        defaults={k: v for k, v in madre_defaults.items() if v is not None}
                    )
                    # --- FIN DE LA CORRECCIÓN --- ### <<<

                request.session.pop('form_data', None)
                
                if 'save_draft' in request.POST:
                    request.session.pop('historia_en_progreso_id', None)
                    messages.success(request, "Borrador de información de padres guardado exitosamente.")
                    redirect_url = reverse('ver_historias')
                    if is_ajax:
                        return JsonResponse({'success': True, 'redirect_url': redirect_url})
                    return redirect(redirect_url)
                    
                messages.success(request, "Información de los padres guardada/actualizada.")

            # ===== INICIO DE LA LÓGICA DE REDIRECCIÓN DINÁMICA (MODIFICADA) =====
                next_step_index = current_step_index
                if next_step_index < len(workflow):
                    next_node_name = workflow[next_step_index]
                    next_view_name = ALL_STEPS[next_node_name]['view_name']
                    # El siguiente paso (ej. 'enfermedad_actual') necesita 'historia_id', 'tipo', y 'objeto_id'.
                    next_kwargs = {
                        'historia_id': historia_id, 
                        'tipo': 'proposito', 
                        'objeto_id': proposito_id
                    }
                    redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                else:
                    redirect_url = reverse('index') # Fallback por si es el último paso
            # ===== FIN DE LA LÓGICA DE REDIRECCIÓN DINÁMICA =====
                
                
                if is_ajax:
                    return JsonResponse({'success': True, 'redirect_url': redirect_url})
                return redirect(redirect_url)

            except Exception as e:
                # El error que recibiste se captura aquí y se muestra
                error_msg = f"Error al guardar información de padres: {e}"
                if is_ajax: return JsonResponse({'success': False, 'errors': {'__all__': [error_msg]}}, status=400)
                messages.error(request, error_msg)
                request.session['form_data'] = request.POST.copy()
                return redirect(request.path_info)
        else:
            if is_ajax: return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            messages.error(request, "No se pudo guardar información de padres. Corrija errores.")
            context['form'] = form # Añadimos el form con errores al contexto
            return render(request, "Padres_proposito.html", context)
    
    
    else: # Lógica para la petición GET (cuando la página carga por primera vez)
            form_data = request.session.pop('form_data', None)
            
            # Preparamos los argumentos para el constructor del formulario.
            # El formulario usará estas instancias para poblarse automáticamente.
            form_kwargs = {
                'padre_instance': padre_instance,
                'madre_instance': madre_instance
            }
            
            if form_data:
                # Si venimos de un POST fallido, usamos los datos de la sesión.
                form = PadresPropositoForm(form_data, **form_kwargs)
            else:
                # Si es una carga normal, instanciamos el formulario pasándole las instancias.
                # No necesitamos construir 'initial' manualmente.
                form = PadresPropositoForm(**form_kwargs)
                if context['editing'] and not form_data: # Usamos la variable 'editing' del contexto
                    messages.info(request, "Editando información de los padres.")

# --- FUERA DEL BLOQUE if/else ---
 # Actualizamos el 'form' en el diccionario de contexto que ya habíamos creado.
    context['form'] = form

# Usamos el 'context' completo para renderizar la plantilla.
# Este return ahora está fuera del bloque 'else' y sirve para ambos casos (GET y POST fallido).
    return render(request, "Padres_proposito.html", context)


@login_required
@genetista_or_admin_required
@never_cache
def crear_antecedentes_personales(request, historia_id, tipo, objeto_id):
    historia = get_object_or_404(HistoriasClinicas, historia_id=historia_id)
    proposito_obj, pareja_obj, context_object_name = None, None, ""
    editing = False
    user_gen_profile = request.user.genetistas
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if tipo == 'proposito':
        proposito_obj = get_object_or_404(Propositos, proposito_id=objeto_id, historia=historia)
        if user_gen_profile.rol == 'GEN' and proposito_obj.historia.genetista != user_gen_profile:
            raise PermissionDenied("No tiene permiso para esta acción.")
        context_object_name = f"{proposito_obj.nombres} {proposito_obj.apellidos}"
        if AntecedentesPersonales.objects.filter(proposito=proposito_obj).exists(): editing = True
    elif tipo == 'pareja':
        pareja_obj = get_object_or_404(Parejas, pareja_id=objeto_id)
        if user_gen_profile.rol == 'GEN':
            p1_hist_gen = pareja_obj.proposito_id_1.historia.genetista if pareja_obj.proposito_id_1.historia else None
            p2_hist_gen = pareja_obj.proposito_id_2.historia.genetista if pareja_obj.proposito_id_2 and pareja_obj.proposito_id_2.historia else None
            if not (p1_hist_gen == user_gen_profile or p2_hist_gen == user_gen_profile):
                 raise PermissionDenied("No tiene permiso para esta acción sobre la pareja.")
        context_object_name = f"Pareja ID: {pareja_obj.pareja_id}"
        if AntecedentesPersonales.objects.filter(pareja=pareja_obj).exists(): editing = True
    else:
        messages.error(request, 'Tipo de objeto no válido.')
        return redirect('index')

    current_node_name = 'antecedentes_personales'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 5

    context = {
        'form': None,
        'historia': historia,
        'tipo': tipo,
        'objeto': proposito_obj or pareja_obj,
        'context_object_name': context_object_name,
        'editing': editing,
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }

    if request.method == 'POST':
        form = AntecedentesDesarrolloNeonatalForm(request.POST)
        if form.is_valid():
            try:
                target_proposito = proposito_obj if tipo == 'proposito' else None
                target_pareja = pareja_obj if tipo == 'pareja' else None
                form.save(proposito=target_proposito, pareja=target_pareja)
                request.session.pop('form_data', None)

                if 'save_draft' in request.POST:
                    request.session.pop('historia_en_progreso_id', None)
                    request.session.pop('historia_workflow', None)
                    messages.success(request, f"Borrador de antecedentes guardado.")
                    redirect_url = reverse('ver_historias')
                else:
                    action_verb = "actualizados" if editing else "guardados"
                    messages.success(request, f"Antecedentes personales {action_verb}.")
                    
                    next_node_name = workflow[current_step_index] if current_step_index < len(workflow) else None
                    if next_node_name:
                        next_view_name = ALL_STEPS[next_node_name]['view_name']
                        next_kwargs = {'historia_id': historia.historia_id, 'tipo': tipo, 'objeto_id': objeto_id}
                        redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                    else:
                        redirect_url = reverse('index')
                
                if is_ajax:
                    return JsonResponse({'success': True, 'redirect_url': redirect_url})
                return redirect(redirect_url)
            except Exception as e:
                error_msg = f'Error al guardar antecedentes: {str(e)}'
                if is_ajax: return JsonResponse({'success': False, 'errors': {'__all__': [error_msg]}}, status=400)
                messages.error(request, error_msg)
                context['form'] = form
                return render(request, 'antecedentes_personales.html', context)
        else:
            if is_ajax: return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            messages.error(request, "No se pudieron guardar los antecedentes. Corrija errores.")
            context['form'] = form
            return render(request, 'antecedentes_personales.html', context)
        
    else: # Lógica GET
        form_data = request.session.pop('form_data', None)
        if form_data:
            form = AntecedentesDesarrolloNeonatalForm(form_data)
        else:
            # ===== INICIO DE LA CORRECCIÓN =====
            initial_data = {}
            if editing:
                messages.info(request, f"Editando antecedentes para {context_object_name}.")
                
                # 1. Definimos el filtro una sola vez para reutilizarlo.
                q_filter = Q(proposito=proposito_obj) if tipo == 'proposito' else Q(pareja=pareja_obj)

                # 2. Buscamos las instancias de los tres modelos relacionados.
                ap_instance = AntecedentesPersonales.objects.filter(q_filter).first()
                dp_instance = DesarrolloPsicomotor.objects.filter(q_filter).first()
                pn_instance = PeriodoNeonatal.objects.filter(q_filter).first()

                # 3. Poblamos el diccionario 'initial_data' con los datos de cada instancia si existen.
                # model_to_dict convierte un objeto de modelo en un diccionario.
                if ap_instance:
                    initial_data.update(model_to_dict(ap_instance))
                if dp_instance:
                    initial_data.update(model_to_dict(dp_instance))
                if pn_instance:
                    initial_data.update(model_to_dict(pn_instance))
            
            # 4. Instanciamos el formulario pasándole el diccionario 'initial_data' ya poblado.
            form = AntecedentesDesarrolloNeonatalForm(initial=initial_data)
            # ===== FIN DE LA CORRECCIÓN =====

    context['form'] = form
    return render(request, 'antecedentes_personales.html', context)

@login_required
@genetista_or_admin_required
@never_cache
def crear_antecedentes_preconcepcionales(request, historia_id, tipo, objeto_id):
    historia = get_object_or_404(HistoriasClinicas, historia_id=historia_id)
    proposito_obj, pareja_obj, context_object_name = None, None, ""
    
    # Declaramos las variables que vamos a usar
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    user_gen_profile = request.user.genetistas
    
    # 1. OBTENER EL OBJETO DE CONTEXTO (Propósito o Pareja)
    if tipo == 'proposito':
        proposito_obj = get_object_or_404(Propositos, proposito_id=objeto_id, historia=historia)
        if user_gen_profile.rol == 'GEN' and proposito_obj.historia.genetista != user_gen_profile:
            raise PermissionDenied("No tiene permiso para esta acción.")
        context_object_name = f"{proposito_obj.nombres} {proposito_obj.apellidos}"
    elif tipo == 'pareja':
        pareja_obj = get_object_or_404(Parejas.objects.select_related('proposito_id_1', 'proposito_id_2'), pareja_id=objeto_id)
        if user_gen_profile.rol == 'GEN':
            p1_gen = pareja_obj.proposito_id_1.historia.genetista if pareja_obj.proposito_id_1.historia else None
            if p1_gen != user_gen_profile:
                raise PermissionDenied("No tiene permiso para esta acción sobre la pareja.")
        context_object_name = f"Pareja: {pareja_obj.proposito_id_1.nombres} y {pareja_obj.proposito_id_2.nombres}"
    else:
        messages.error(request, "Tipo de objeto no válido.")
        return redirect('index')

    # 2. OBTENER O CREAR LA INSTANCIA DE ANTECEDENTES ASOCIADA
    # Esta es la corrección clave.
    if tipo == 'proposito':
        # Busca un antecedente para este propósito. Si no existe, lo crea.
        instance_to_edit, created = AntecedentesFamiliaresPreconcepcionales.objects.get_or_create(
            proposito=proposito_obj
        )
    else: # tipo == 'pareja'
        # Busca un antecedente para esta pareja. Si no existe, lo crea.
        instance_to_edit, created = AntecedentesFamiliaresPreconcepcionales.objects.get_or_create(
            pareja=pareja_obj
        )

    # El modo de edición es verdadero si la instancia ya existía (no fue 'created')
    editing = not created

    # 3. PREPARAR EL CONTEXTO PARA EL STEPPER Y LA PLANTILLA
    current_node_name = 'antecedentes_preconcepcionales'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 6

    context = {
        'form': None,
        'historia': historia,
        'tipo': tipo,
        'objeto': proposito_obj or pareja_obj,
        'context_object_name': context_object_name,
        'editing': editing,
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }

    # ... a partir de aquí comienza tu `if request.method == 'POST':` ...

    # 3. MANEJO DE LA PETICIÓN
    if request.method == 'POST':
        form = AntecedentesPreconcepcionalesForm(request.POST, instance=instance_to_edit)
        if form.is_valid():
            try:
                form.save()
                request.session.pop('form_data', None)

                redirect_url = None

                # LÓGICA DE REDIRECCIÓN BASADA EN EL BOTÓN PRESIONADO
                if 'save_draft' in request.POST:
                    request.session.pop('historia_en_progreso_id', None)
                    request.session.pop('historia_workflow', None)
                    messages.success(request, "Borrador de antecedentes preconcepcionales guardado.")
                    redirect_url = reverse('ver_historias')

                elif 'save_and_exam_proposito' in request.POST and proposito_obj:
                    messages.success(request, "Datos guardados. Procediendo a Examen Físico.")
                    redirect_url = reverse('examen_fisico_crear_editar', kwargs={'proposito_id': proposito_obj.proposito_id})

                elif 'save_and_exam_p1' in request.POST and pareja_obj:
                    messages.success(request, f"Datos guardados. Procediendo a Examen Físico para {pareja_obj.proposito_id_1.nombres}.")
                    redirect_url = reverse('examen_fisico_crear_editar', kwargs={'proposito_id': pareja_obj.proposito_id_1.proposito_id}) + f"?pareja_id={pareja_obj.pareja_id}"

                elif 'save_and_exam_p2' in request.POST and pareja_obj:
                    messages.success(request, f"Datos guardados. Procediendo a Examen Físico para {pareja_obj.proposito_id_2.nombres}.")
                    redirect_url = reverse('examen_fisico_crear_editar', kwargs={'proposito_id': pareja_obj.proposito_id_2.proposito_id}) + f"?pareja_id={pareja_obj.pareja_id}"

                elif 'save_and_continue' in request.POST:
                    messages.success(request, "Datos guardados. Procediendo a Evaluación Genética.")
                    try:
                        evaluacion_view_name = ALL_STEPS['evaluacion']['view_name']
                        evaluacion_kwargs = {'historia_id': historia_id, 'tipo': tipo, 'objeto_id': objeto_id}
                        redirect_url = reverse(evaluacion_view_name, kwargs=evaluacion_kwargs)
                    except KeyError:
                        messages.error(request, "Error: El paso de 'Evaluación' no está definido en el flujo.")
                        redirect_url = reverse('index')
                
                if not redirect_url:
                    messages.warning(request, "Acción no reconocida, redirigiendo a la página de inicio.")
                    redirect_url = reverse('index')

                if is_ajax:
                    return JsonResponse({'success': True, 'redirect_url': redirect_url})
                return redirect(redirect_url)

            except Exception as e:
                error_msg = f'Error al guardar antecedentes preconcepcionales: {str(e)}'
                if is_ajax: return JsonResponse({'success': False, 'errors': {'__all__': [error_msg]}}, status=400)
                messages.error(request, error_msg)
                context['form'] = form
                return render(request, 'antecedentes_preconcepcionales.html', context)
        else: # Formulario no es válido
            if is_ajax: return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            messages.error(request, "No se pudieron guardar los antecedentes. Por favor, corrija los errores.")
            context['form'] = form
            return render(request, 'antecedentes_preconcepcionales.html', context)
    
    else: # Lógica GET
        form_data = request.session.pop('form_data', None)
        form = AntecedentesPreconcepcionalesForm(form_data or None, initial=vars(instance_to_edit) if instance_to_edit and not form_data else None)
        if instance_to_edit and not form_data: messages.info(request, f"Editando antec. preconcepcionales para {context_object_name}.")


    context['form'] = form
    return render(request, 'antecedentes_preconcepcionales.html', context)


# En tu archivo views.py

@login_required
@genetista_or_admin_required
@never_cache
def crear_examen_fisico(request, proposito_id):
    proposito = get_object_or_404(Propositos, pk=proposito_id)
    user_profile = request.user.genetistas
    if user_profile.rol == 'GEN' and proposito.historia.genetista != user_profile:
        raise PermissionDenied("No tiene permiso para modificar el examen físico de este propósito.")
    
    examen_existente = ExamenFisico.objects.filter(proposito=proposito).first()
    editing = bool(examen_existente)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    
    # Lógica para encontrar si hay otro miembro de la pareja con examen pendiente
    pareja = None
    otro_proposito_pendiente = None 

    pareja_id = request.GET.get('pareja_id') or request.POST.get('pareja_id')
    if pareja_id:
        pareja = get_object_or_404(Parejas.objects.select_related('proposito_id_1', 'proposito_id_2'), pk=pareja_id)
        otro_proposito_obj = pareja.proposito_id_2 if proposito.pk == pareja.proposito_id_1.pk else pareja.proposito_id_1
        
        if otro_proposito_obj and not ExamenFisico.objects.filter(proposito=otro_proposito_obj).exists():
             otro_proposito_pendiente = otro_proposito_obj

    # Lógica de contexto del Stepper
    current_node_name = 'examen_fisico'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 7 # Fallback razonable
    
    context = {
        'form': None, 
        'proposito': proposito, 
        'editing': editing,
        'pareja': pareja, 
        'otro_proposito_pendiente': otro_proposito_pendiente,
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }

    if request.method == 'POST':
        form = ExamenFisicoForm(request.POST, instance=examen_existente)
        form.proposito_instance = proposito
        if form.is_valid():
            form.save()
            request.session.pop('form_data', None)

            redirect_url = None
            if 'save_draft' in request.POST:
                request.session.pop('historia_en_progreso_id', None)
                request.session.pop('historia_workflow', None)
                messages.success(request, f"Borrador del examen físico para {proposito.nombres} guardado.")
                redirect_url = reverse('ver_historias')
            else:
                action_verb = "actualizado" if examen_existente else "guardado"
                messages.success(request, f"Examen físico para {proposito.nombres} {action_verb}.")
                
                # Lógica de redirección basada en el botón presionado
                if 'save_and_go_to_other' in request.POST and otro_proposito_pendiente:
                    messages.info(request, f"Ahora puede completar el examen para {otro_proposito_pendiente.nombres}.")
                    redirect_url = reverse('examen_fisico_crear_editar', kwargs={'proposito_id': otro_proposito_pendiente.proposito_id}) + f"?pareja_id={pareja.pareja_id}"
                else:
                    # Redirección normal al siguiente paso del workflow
                    next_node_name = workflow[current_step_index] if current_step_index < len(workflow) else None
                    if next_node_name:
                        next_view_name = ALL_STEPS[next_node_name]['view_name']
                        context_tipo = 'pareja' if pareja else 'proposito'
                        context_objeto_id = pareja.pareja_id if pareja else proposito.proposito_id
                        next_kwargs = {'historia_id': proposito.historia.historia_id, 'tipo': context_tipo, 'objeto_id': context_objeto_id}
                        redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                    else:
                        # Si no hay siguiente paso, redirige a la finalización o al índice
                        redirect_url = reverse('index')
            
            if is_ajax:
                return JsonResponse({'success': True, 'redirect_url': redirect_url})
            return redirect(redirect_url)
        else:
            # ===== INICIO DE LA CORRECCIÓN =====
            # Si la petición es AJAX y el formulario es inválido, devolvemos los errores como JSON.
            if is_ajax:
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            
            # Si no es AJAX, continuamos con el comportamiento normal: mostrar mensajes y renderizar la página.
            messages.error(request, f"No se pudo guardar el examen físico. Por favor, corrija los errores.")
            context['form'] = form
            return render(request, 'examen_fisico.html', context)
            # ===== FIN DE LA CORRECCIÓN =====
        
    else: # Lógica GET
        form = ExamenFisicoForm(instance=examen_existente)
        if editing:
            messages.info(request, f"Editando examen físico para {proposito.nombres}.")

    context['form'] = form
    return render(request, 'examen_fisico.html', context)

@login_required
@genetista_or_admin_required
@never_cache
@transaction.atomic
def diagnosticos_plan_estudio(request, historia_id, tipo, objeto_id):
    try:
        if tipo == 'proposito':
            parent_object = get_object_or_404(Propositos, pk=objeto_id)
            lookup_kwargs = {'proposito': parent_object}
        elif tipo == 'pareja':
            parent_object = get_object_or_404(Parejas, pk=objeto_id)
            lookup_kwargs = {'pareja': parent_object}
        else:
            messages.error(request, "Tipo de objeto no válido.")
            return redirect('index')
    except Exception as e:
        messages.error(request, f"No se pudo encontrar el objeto de referencia: {e}")
        return redirect('index')

    evaluacion_instance, created = EvaluacionGenetica.objects.get_or_create(**lookup_kwargs, defaults={'signos_clinicos': ''})


    # ===== INICIO DE LA LÓGICA DE CONTEXTO DINÁMICO (MODIFICADA) =====
    current_node_name = 'evaluacion'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 8 # Valor por defecto razonable

    context = {
        'form': None, 
        'diagnostico_formset': None,
        'plan_formset': None,
        'parent_object': parent_object, 
        'historia_id': historia_id,
        'is_new_consultation': False,
        # --- NUEVAS VARIABLES PARA EL STEPPER ---
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }
    # ===== FIN DE LA LÓGICA DE CONTEXTO DINÁMICO =====



    if request.method == 'POST':
        form = EvaluacionGeneticaForm(request.POST, instance=evaluacion_instance)
        diagnostico_formset = DiagnosticoFormSet(request.POST, instance=evaluacion_instance, prefix='diagnostico')
        plan_formset = PlanEstudioFormSet(request.POST, instance=evaluacion_instance, prefix='plan')
        
        if form.is_valid() and diagnostico_formset.is_valid() and plan_formset.is_valid():
            form.save()
            diagnostico_formset.save()
            plan_formset.save()


         # --- INICIO: LÓGICA DE REDIRECCIÓN DINÁMICA ---
            if 'save_draft' in request.POST:
                request.session.pop('historia_en_progreso_id', None)
                request.session.pop('historia_workflow', None)
                messages.success(request, "Borrador de evaluación genética guardado exitosamente.")
                redirect_url = reverse('ver_historias')
            else:
                messages.success(request, '¡Evaluación genética guardada exitosamente!')
                
                next_node_name = workflow[current_step_index] if current_step_index < len(workflow) else None
                if next_node_name:
                    next_view_name = ALL_STEPS[next_node_name]['view_name']
                    # El siguiente paso (genealogia) necesita los mismos kwargs
                    next_kwargs = {'historia_id': historia_id, 'tipo': tipo, 'objeto_id': objeto_id}
                    redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                else:
                    redirect_url = reverse('index') # O 'index'

         # --- FIN: LÓGICA DE REDIRECCIÓN DINÁMICA ---

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect_url': redirect_url})
            return redirect(redirect_url)
        
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                errors = {}
                errors.update(form.errors.get_json_data())
                for i, form_errors in enumerate(diagnostico_formset.errors):
                    if form_errors:
                        for field, error_list in form_errors.items(): errors[f'diagnostico-{i}-{field}'] = error_list
                for i, form_errors in enumerate(plan_formset.errors):
                    if form_errors:
                        for field, error_list in form_errors.items(): errors[f'plan-{i}-{field}'] = error_list
                if diagnostico_formset.non_form_errors(): errors['diagnostico-non-form'] = diagnostico_formset.non_form_errors()
                if plan_formset.non_form_errors(): errors['plan-non-form'] = plan_formset.non_form_errors()
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            
            messages.error(request, 'Por favor, corrija los errores en el formulario.')

    else: # Lógica GET
        form = EvaluacionGeneticaForm(instance=evaluacion_instance)
        diagnostico_formset = DiagnosticoFormSet(instance=evaluacion_instance, prefix='diagnostico')
        plan_formset = PlanEstudioFormSet(instance=evaluacion_instance, prefix='plan')

    # Actualizamos el contexto con los formularios y renderizamos
    context.update({
        'form': form,
        'diagnostico_formset': diagnostico_formset,
        'plan_formset': plan_formset,
    })
    return render(request, 'diagnosticos_plan.html', context)



@login_required
@genetista_or_admin_required
@never_cache
@transaction.atomic
def nueva_consulta_view(request, evaluacion_id):
    """
    Vista para añadir una nueva consulta (diagnósticos y planes) a una evaluación existente.
    Reutiliza la plantilla 'diagnosticos_plan.html' pero oculta la sección de signos clínicos.
    """
    evaluacion_instance = get_object_or_404(EvaluacionGenetica, pk=evaluacion_id)
    
    # Verificación de permisos
    user_profile = request.user.genetistas
    if user_profile.rol == 'GEN':
        genetista_historia = None
        if evaluacion_instance.proposito:
            genetista_historia = evaluacion_instance.proposito.historia.genetista
        elif evaluacion_instance.pareja:
            genetista_historia = evaluacion_instance.pareja.proposito_id_1.historia.genetista
        if genetista_historia != user_profile:
            raise PermissionDenied("No tiene permiso para añadir consultas a esta evaluación.")
            
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method == 'POST':
        # No necesitamos el form principal, solo los formsets
        diagnostico_formset = DiagnosticoFormSet(request.POST, instance=evaluacion_instance, prefix='diagnostico')
        plan_formset = PlanEstudioFormSet(request.POST, instance=evaluacion_instance, prefix='plan')

        if diagnostico_formset.is_valid() and plan_formset.is_valid():
            diagnostico_formset.save()
            plan_formset.save()
            
            messages.success(request, 'Nueva consulta añadida exitosamente.')
            redirect_url = reverse('gestion_pacientes') # Redirigir de vuelta a la gestión

            if is_ajax:
                return JsonResponse({'success': True, 'redirect_url': redirect_url})
            return redirect(redirect_url)
        else:
            if is_ajax:
                errors = {}
                # Agregamos los errores de los formsets al diccionario de errores
                for i, form_errors in enumerate(diagnostico_formset.errors):
                    if form_errors:
                        for field, error_list in form_errors.items(): errors[f'diagnostico-{i}-{field}'] = error_list
                for i, form_errors in enumerate(plan_formset.errors):
                    if form_errors:
                        for field, error_list in form_errors.items(): errors[f'plan-{i}-{field}'] = error_list
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            
            messages.error(request, 'Por favor, corrija los errores en el formulario.')

    # Lógica GET
    # Para que los formsets aparezcan vacíos (solo con 'extra=1'),
    # les pasamos un queryset vacío. Esto evita que carguen los datos existentes.
    diagnostico_formset = DiagnosticoFormSet(
        instance=evaluacion_instance, 
        prefix='diagnostico', 
        queryset=DiagnosticoPresuntivo.objects.none()
    )
    plan_formset = PlanEstudioFormSet(
        instance=evaluacion_instance, 
        prefix='plan',
        queryset=PlanEstudio.objects.none()
    )
    
    # El `form` principal de EvaluacionGenetica no se usa, pero lo pasamos para que la plantilla no falle
    form_dummy = EvaluacionGeneticaForm()

    context = {
        'form': form_dummy,
        'diagnostico_formset': diagnostico_formset,
        'plan_formset': plan_formset,
        'historia_id': evaluacion_instance.proposito.historia.historia_id if evaluacion_instance.proposito else evaluacion_instance.pareja.proposito_id_1.historia.historia_id,
        'is_new_consultation': True # La variable clave para la plantilla
    }
    return render(request, 'diagnosticos_plan.html', context)

@login_required
@genetista_or_admin_required
@never_cache
@transaction.atomic
def autorizaciones_view(request, historia_id, tipo, objeto_id):
    historia = get_object_or_404(HistoriasClinicas, pk=historia_id)
    propositos_a_procesar = []
    
    if tipo == 'proposito':
        proposito = get_object_or_404(Propositos, pk=objeto_id)
        propositos_a_procesar.append(proposito)
    elif tipo == 'pareja':
        pareja = get_object_or_404(Parejas.objects.select_related('proposito_id_1', 'proposito_id_2'), pk=objeto_id)
        propositos_a_procesar.append(pareja.proposito_id_1)
        propositos_a_procesar.append(pareja.proposito_id_2)
    else:
        messages.error(request, "Contexto no válido.")
        return redirect('index')


    # ===== INICIO DE LA LÓGICA DE CONTEXTO DINÁMICO =====
    current_node_name = 'autorizacion'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 10 # Valor por defecto razonable

    context = {
        'autorizacion_contexts': [], # Se llenará más adelante
        'historia': historia,
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }
    # ===== FIN DE LA LÓGICA DE CONTEXTO DINÁMICO =====


    if request.method == 'POST':
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        all_forms_valid = True
        errors = {}

        for proposito in propositos_a_procesar:
            prefix = f'form_{proposito.proposito_id}'
            autorizacion_instance, _ = Autorizaciones.objects.get_or_create(proposito=proposito)
            form = AutorizacionForm(request.POST, instance=autorizacion_instance, prefix=prefix, proposito=proposito)

            if form.is_valid():
                autorizacion = form.save(commit=False)
                autorizacion.proposito = proposito

                # ===== INICIO DE LA LÓGICA CORREGIDA PARA PROCESAR LA FIRMA =====
                signature_base64 = form.cleaned_data.get('signature_data')
                
                # Solo procesamos si el campo de firma no está vacío (es decir, si el usuario dibujó una nueva firma).
                if signature_base64:
                    try:
                        # Formato esperado: "data:image/png;base64,iVBORw0KGgo..."
                        format, imgstr = signature_base64.split(';base64,') 
                        ext = format.split('/')[-1] # ej. 'png'
                        file_name = f'firma_{proposito.identificacion}_{timezone.now().strftime("%Y%m%d%H%M%S")}.{ext}'
                        data = ContentFile(base64.b64decode(imgstr), name=file_name)
                        
                        # Si ya existía una firma, la eliminamos del almacenamiento antes de guardar la nueva.
                        if autorizacion.pk and autorizacion.archivo_autorizacion:
                            autorizacion.archivo_autorizacion.delete(save=False)
                            
                        autorizacion.archivo_autorizacion = data
                    except (ValueError, TypeError, IndexError):
                        # Este error ocurre si la cadena base64 no tiene el formato esperado.
                        form.add_error(None, f"Formato de firma inválido para {proposito.nombres}.")
                        all_forms_valid = False
                        errors[f'{prefix}-__all__'] = [{'message': 'Formato de firma inválido.', 'code': ''}]

                if all_forms_valid:
                    autorizacion.save()
            else:
                all_forms_valid = False
                for field, error_list in form.errors.items():
                    errors[f"{prefix}-{field}"] = error_list

        if all_forms_valid:
            historia.estado = HistoriasClinicas.ESTADO_FINALIZADA
            historia.save(update_fields=['estado'])
            
            messages.success(request, "¡Historia clínica finalizada y guardada exitosamente!")
            
            request.session.pop('historia_en_progreso_id', None)
            request.session.pop('source_is_edit_list', None)
            request.session.pop('historia_workflow', None)

             # ===== INICIO DE LA LÓGICA DE REDIRECCIÓN DINÁMICA (MODIFICADA) =====
            # Como este es el último paso, siempre redirige a la misma vista de finalización.
            # No necesita consultar el workflow.
            redirect_url = reverse('index') # O 'index' o 'ver_historias' si lo prefieres
            # ===== FIN DE LA LÓGICA DE REDIRECCIÓN DINÁMICA =====

            if is_ajax:
                return JsonResponse({'success': True, 'redirect_url': redirect_url})
            return redirect(redirect_url)
        else:
            if is_ajax:
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            messages.error(request, "Por favor, corrija los errores en el formulario.")

    # Lógica GET
    autorizacion_contexts = []
    for proposito in propositos_a_procesar:
        instance, _ = Autorizaciones.objects.get_or_create(proposito=proposito)
        form = AutorizacionForm(instance=instance, proposito=proposito, prefix=f'form_{proposito.proposito_id}')
        context_item = {
            'proposito': proposito,
            'form': form,
            'is_minor': proposito.is_minor(),
        }
        autorizacion_contexts.append(context_item)

    # Actualizamos el 'autorizacion_contexts' en el diccionario de contexto principal
    context['autorizacion_contexts'] = autorizacion_contexts
    
    return render(request, 'autorizaciones.html', context)


@require_POST # Esta vista solo debe aceptar peticiones POST
@login_required
@admin_required
def edit_user_admin(request, user_id):
    user_to_edit = get_object_or_404(User, pk=user_id)
    form = AdminUserEditForm(request.POST, instance=user_to_edit)

    if form.is_valid():
        try:
            with transaction.atomic():
                form.save()
            messages.success(request, f"Usuario '{user_to_edit.username}' actualizado exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al actualizar el usuario: {e}")
    else:
        # Si el formulario no es válido, creamos un mensaje de error detallado
        error_list = []
        for field, errors in form.errors.items():
            field_name = form.fields[field].label or field
            error_list.append(f"{field_name}: {', '.join(errors)}")
        error_message = "No se pudo actualizar el usuario. Errores: " + "; ".join(error_list)
        messages.error(request, error_message)

    return redirect('gestion_usuarios')

# ====== FIN DEL CÓDIGO A AÑADIR EN views.py ======





@login_required
@all_roles_required
def ver_proposito(request, proposito_id):
    proposito = get_object_or_404(Propositos, pk=proposito_id)
    user_gen_profile = request.user.genetistas

    if user_gen_profile.rol == 'GEN' and proposito.historia.genetista != user_gen_profile:
        raise PermissionDenied("No tiene permiso para ver este propósito.")
    elif user_gen_profile.rol == 'LEC':
        if not user_gen_profile.associated_genetista or \
           proposito.historia.genetista != user_gen_profile.associated_genetista:
            raise PermissionDenied("No tiene permiso para ver este propósito.")

    examen_fisico = ExamenFisico.objects.filter(proposito=proposito).first()
    padres_info = InformacionPadres.objects.filter(proposito=proposito)
    antecedentes_personales = AntecedentesPersonales.objects.filter(proposito=proposito).first()
    desarrollo_psicomotor = DesarrolloPsicomotor.objects.filter(proposito=proposito).first()
    periodo_neonatal = PeriodoNeonatal.objects.filter(proposito=proposito).first()
    antecedentes_familiares = AntecedentesFamiliaresPreconcepcionales.objects.filter(proposito=proposito).first()
    evaluacion_genetica = EvaluacionGenetica.objects.filter(proposito=proposito).first()
    diagnosticos, planes_estudio = [], []
    if evaluacion_genetica:
        diagnosticos = DiagnosticoPresuntivo.objects.filter(evaluacion=evaluacion_genetica).order_by('orden')
        planes_estudio = PlanEstudio.objects.filter(evaluacion=evaluacion_genetica).order_by('pk')

    return render(request, "ver_proposito.html", {
        'proposito': proposito, 'examen_fisico': examen_fisico,
        'padres_info': {p.tipo: p for p in padres_info},
        'antecedentes_personales': antecedentes_personales,
        'desarrollo_psicomotor': desarrollo_psicomotor, 'periodo_neonatal': periodo_neonatal,
        'antecedentes_familiares': antecedentes_familiares, 'evaluacion_genetica': evaluacion_genetica,
        'diagnosticos_presuntivos': diagnosticos, 'planes_estudio': planes_estudio,
    })
    



# views.py

@login_required
@admin_required
@never_cache
def gestion_usuarios_view(request):
    clear_editing_session(request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method == 'POST':
        # ... (la lógica del POST para crear usuario se queda igual) ...
        if 'create_user_submit' in request.POST:
            form = AdminUserCreationForm(request.POST)
            if form.is_valid():
                try:
                    form.save()
                    request.session.pop('form_data', None)
                    messages.success(request, "Usuario creado exitosamente.")
                    redirect_url = reverse('gestion_usuarios')
                    if is_ajax:
                        return JsonResponse({'success': True, 'redirect_url': redirect_url})
                    return redirect(redirect_url)
                except Exception as e:
                    error_msg = f"Error al crear usuario: {e}"
                    if is_ajax:
                        return JsonResponse({'success': False, 'errors': {'__all__': [error_msg]}}, status=400)
                    messages.error(request, error_msg)
                    request.session['form_data'] = request.POST.copy() # <<< PRG FIX
                    return redirect(request.path_info) # <<< PRG FIX
            else: # Form is invalid
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': form.errors}, status=400)
                messages.error(request, "Error al crear usuario. Por favor, corrija los errores.")
                request.session['form_data'] = request.POST.copy() # <<< PRG FIX
                return redirect(request.path_info) # <<< PRG FIX
    
    # GET request logic
    form_data = request.session.pop('form_data', None) # <<< PRG FIX
    user_creation_form_instance = AdminUserCreationForm(form_data) if form_data else AdminUserCreationForm()
    
    # ====== INICIO DEL CÓDIGO A AÑADIR EN views.py ======
    # Creamos una instancia del formulario de edición para usar en el modal de edición.
    # Esto nos permite tener campos pre-configurados con los IDs y clases correctos.
    user_edit_form_instance = AdminUserEditForm()
    # ====== FIN DEL CÓDIGO A AÑADIR EN views.py ======

    total_users_count = User.objects.count()
    active_users_count = User.objects.filter(is_active=True).count()
    role_counts = Genetistas.objects.aggregate(
        admin_count=Count('user_id', filter=Q(rol='ADM')),
        genetista_count=Count('user_id', filter=Q(rol='GEN')),
        lector_count=Count('user_id', filter=Q(rol='LEC')),
    )
    users_qs = User.objects.select_related('genetistas').all().order_by('last_name', 'first_name')
    search_query = request.GET.get('buscar-usuario', '').strip()
    role_filter = request.GET.get('role_filter', '').strip()

    if search_query:
        users_qs = users_qs.filter(Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query) | Q(email__icontains=search_query))
    if role_filter:
        users_qs = users_qs.filter(genetistas__rol=role_filter)

    context = {
        'user_creation_form': user_creation_form_instance,
        # ====== INICIO DE LA LÍNEA A AÑADIR EN views.py ======
        'user_edit_form': user_edit_form_instance, # Pasamos el nuevo formulario al contexto
        # ====== FIN DE LA LÍNEA A AÑADIR EN views.py ======
        'total_users': total_users_count, 'active_users': active_users_count,
        'admin_users_count': role_counts['admin_count'], 'genetista_users_count': role_counts['genetista_count'], 'lector_users_count': role_counts['lector_count'],
        'users_list': users_qs, 'search_query': search_query, 'current_role_filter': role_filter,
        'genetista_roles_for_filter': Genetistas.ROL_CHOICES,
        'form_errors_exist': bool(user_creation_form_instance.errors)
    }
    return render(request, 'gestion_usuarios.html', context)


@login_required
@all_roles_required
@never_cache
def ver_historias(request):
    """
    Vista mejorada para listar, filtrar y gestionar historias clínicas con control de acceso basado en roles.
    """
    # ... (lógica de clear_editing_session y permisos sin cambios) ...
    clear_editing_session(request)

    try:
        user_profile = request.user.genetistas
    except Genetistas.DoesNotExist:
        messages.error(request, "Su perfil de usuario no está configurado.")
        return redirect('index')

    base_qs = HistoriasClinicas.objects.select_related('genetista__user').prefetch_related('propositos_set').order_by('-fecha_ingreso')

    if user_profile.rol in ['GEN', 'LEC']:
        base_qs = base_qs.exclude(estado=HistoriasClinicas.ESTADO_ARCHIVADA)

    if user_profile.rol == 'GEN':
        historias_qs = base_qs.filter(genetista=user_profile)
    elif user_profile.rol == 'LEC':
        if user_profile.associated_genetista:
            historias_qs = base_qs.filter(genetista=user_profile.associated_genetista)
        else:
            historias_qs = HistoriasClinicas.objects.none()
            messages.warning(request, "Usted es un Lector sin genetista asociado y no puede ver historias.")
    else: # ADM o Superuser
        historias_qs = base_qs
    
    # ... (lógica de filtros de búsqueda sin cambios) ...
    search_query = request.GET.get('buscar_historia', '').strip()
    estado_query = request.GET.get('estado_historia', '').strip()
    genetista_query_id = request.GET.get('genetista', '').strip()
    motivo_query = request.GET.get('motivo_consulta', '').strip()

    if search_query:
        historias_qs = historias_qs.filter(
            Q(numero_historia__icontains=search_query) |
            Q(propositos__nombres__icontains=search_query) |
            Q(propositos__apellidos__icontains=search_query)
        ).distinct()

    if estado_query and estado_query != 'todos':
        historias_qs = historias_qs.filter(estado=estado_query)
    
    if (user_profile.rol == 'ADM' or request.user.is_superuser) and genetista_query_id and genetista_query_id != 'todos':
        historias_qs = historias_qs.filter(genetista_id=genetista_query_id)

    if motivo_query and motivo_query != 'todos':
        historias_qs = historias_qs.filter(motivo_tipo_consulta=motivo_query)

    # --- INICIO DE LA MODIFICACIÓN ---
    # MODIFICACIÓN: Incluimos GEN y ADM en la lista para el filtro del template
    genetistas_list_for_filter = Genetistas.objects.filter(
        rol__in=['GEN', 'ADM']
    ).select_related('user').order_by('user__first_name')
    # --- FIN DE LA MODIFICACIÓN ---

    context = {
        'historias_list': historias_qs,
        'genetistas_list': genetistas_list_for_filter,
        'estado_choices': [c for c in HistoriasClinicas.ESTADO_CHOICES if c[0] != HistoriasClinicas.ESTADO_ARCHIVADA] if user_profile.rol != 'ADM' else HistoriasClinicas.ESTADO_CHOICES,
        'motivo_choices': HistoriasClinicas._meta.get_field('motivo_tipo_consulta').choices,
        'archivar_form': ArchivarHistoriaForm()
    }
    return render(request, 'ver_historias.html', context)


# En tu archivo views.py

@login_required
@genetista_or_admin_required
@never_cache
def genealogia_view(request, historia_id, tipo, objeto_id):
    """
    Gestiona la carga y edición de la imagen de la genealogía.
    """
    historia = get_object_or_404(HistoriasClinicas, historia_id=historia_id)
    context_object = None
    context_object_name = ""
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    # --- Lógica para obtener el contexto del objeto (sin cambios, ya está bien) ---
    if tipo == 'proposito':
        context_object = get_object_or_404(Propositos, proposito_id=objeto_id, historia=historia)
        context_object_name = f"{context_object.nombres} {context_object.apellidos}"
    elif tipo == 'pareja':
        pareja = get_object_or_404(Parejas.objects.select_related('proposito_id_1', 'proposito_id_2'), pareja_id=objeto_id)
        if pareja.proposito_id_1.historia == historia or (pareja.proposito_id_2 and pareja.proposito_id_2.historia == historia):
            context_object = pareja
            p1_nombre = pareja.proposito_id_1.nombres if pareja.proposito_id_1 else "N/A"
            p2_nombre = pareja.proposito_id_2.nombres if pareja.proposito_id_2 else "N/A"
            context_object_name = f"Pareja: {p1_nombre} y {p2_nombre}"
        else:
            messages.error(request, 'La pareja no corresponde a esta historia clínica.')
            return redirect('index')

    if not context_object:
        messages.error(request, 'Contexto no válido.')
        return redirect('index')
        
    antecedente, created = AntecedentesFamiliaresPreconcepcionales.objects.get_or_create(
        **{tipo: context_object}
    )
    editing = bool(antecedente.genealogia_foto)

    # ===== INICIO: CONTEXTO DINÁMICO (DEFINIDO UNA SOLA VEZ) =====
    current_node_name = 'genealogia'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        current_step_index = 9

    context = {
        'form': None, # Se llenará más adelante
        'historia': historia,
        'context_object_name': context_object_name,
        'editing': editing,
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }
    # ===== FIN: CONTEXTO DINÁMICO =====

    if request.method == 'POST':
        form = GenealogiaForm(request.POST, request.FILES, instance=antecedente)
        if form.is_valid():
            # Tu lógica para manejar la eliminación de la foto es correcta
            if request.POST.get(f'{form.prefix}-genealogia_foto-clear') and not request.FILES.get(f'{form.prefix}-genealogia_foto'):
                antecedente.genealogia_foto.delete(save=False)
                antecedente.genealogia_foto = None

            form.save()

            # ===== INICIO: LÓGICA DE REDIRECCIÓN DINÁMICA (CORREGIDA) =====
            if 'save_draft' in request.POST:
                request.session.pop('historia_en_progreso_id', None)
                request.session.pop('historia_workflow', None)
                messages.success(request, "Borrador de genealogía guardado.")
                redirect_url = reverse('ver_historias')
            else:
                messages.success(request, "Genealogía guardada exitosamente.")
                next_step_index = current_step_index
                if next_step_index < len(workflow):
                    next_node_name = workflow[next_step_index]
                    next_view_name = ALL_STEPS[next_node_name]['view_name']
                    # El siguiente paso, 'autorizaciones_crear', necesita los mismos kwargs
                    next_kwargs = {'historia_id': historia_id, 'tipo': tipo, 'objeto_id': objeto_id}
                    redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                else:
                    redirect_url = reverse('flow_completion') # O 'index'
            # ===== FIN: LÓGICA DE REDIRECCIÓN DINÁMICA =====
            
            # La respuesta AJAX no cambia
            if is_ajax:
                return JsonResponse({'success': True, 'redirect_url': redirect_url})
            return redirect(redirect_url)
        else:
            if is_ajax:
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            
            # Si el POST no es AJAX y falla, renderizamos de nuevo con el contexto completo
            messages.error(request, "No se pudo guardar la genealogía. Corrija los errores.")
            context['form'] = form # Añadimos el formulario con errores
            return render(request, 'genealogia.html', context)

    # Lógica GET
    form = GenealogiaForm(instance=antecedente)
    context['form'] = form # Actualizamos la clave 'form' en nuestro contexto
    return render(request, 'genealogia.html', context)


@login_required
@genetista_or_admin_required
@never_cache
def enfermedad_actual(request, historia_id, tipo, objeto_id):
    historia = get_object_or_404(HistoriasClinicas, historia_id=historia_id)
    proposito_obj, pareja_obj, context_object_name = None, None, ""
    antecedentes_instance = None
    user_gen_profile = request.user.genetistas
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if tipo == 'proposito':
        proposito_obj = get_object_or_404(Propositos, proposito_id=objeto_id, historia=historia)
        if user_gen_profile.rol == 'GEN' and proposito_obj.historia.genetista != user_gen_profile:
            raise PermissionDenied("No tiene permiso para esta acción.")
        context_object_name = f"{proposito_obj.nombres} {proposito_obj.apellidos}"
        lookup_kwargs = {'proposito': proposito_obj}
    elif tipo == 'pareja':
        pareja_obj = get_object_or_404(Parejas, pareja_id=objeto_id)
        if user_gen_profile.rol == 'GEN' and pareja_obj.proposito_id_1.historia.genetista != user_gen_profile:
            raise PermissionDenied("No tiene permiso para esta acción sobre la pareja.")
        context_object_name = f"Pareja: {pareja_obj.proposito_id_1.nombres} y {pareja_obj.proposito_id_2.nombres}"
        lookup_kwargs = {'pareja': pareja_obj}
    else:
        messages.error(request, 'Tipo de objeto no válido.')
        return redirect('index')

    antecedentes_instance, created = AntecedentesPersonales.objects.get_or_create(**lookup_kwargs)
    editing = bool(antecedentes_instance.enfermedad_actual)


    # ===== INICIO DE LA LÓGICA DE CONTEXTO DINÁMICO (MODIFICADA) =====
    current_node_name = 'enfermedad_actual'
    workflow = request.session.get('historia_workflow', [])
    try:
        current_step_index = workflow.index(current_node_name) + 1
    except (ValueError, TypeError):
        # Asignamos un valor por defecto razonable si falla
        current_step_index = 4 if tipo == 'proposito' else 3

    context = {
        'form': None, 
        'historia': historia, 
        'tipo': tipo, 
        'objeto': proposito_obj or pareja_obj, 
        'context_object_name': context_object_name, 
        'editing': editing,
        # --- NUEVAS VARIABLES PARA EL STEPPER ---
        'current_node': current_node_name,
        'current_step_index': current_step_index,
        'ALL_STEPS': ALL_STEPS
    }
    # ===== FIN DE LA LÓGICA DE CONTEXTO DINÁMICO =====


    if request.method == 'POST':
        form = EnfermedadActualForm(request.POST, instance=antecedentes_instance)
        if form.is_valid():
            try:
                form.save()
                request.session.pop('form_data', None)

                if 'save_draft' in request.POST:
                    request.session.pop('historia_en_progreso_id', None)
                    messages.success(request, f"Borrador de enfermedad actual para {context_object_name} guardado.")
                    redirect_url = reverse('ver_historias')
                    if is_ajax:
                        return JsonResponse({'success': True, 'redirect_url': redirect_url})
                    return redirect(redirect_url)

                action_verb = "actualizada" if editing else "guardada"
                messages.success(request, f"Enfermedad actual {action_verb} para {context_object_name}.")
                
                
                 # ===== INICIO DE LA LÓGICA DE REDIRECCIÓN DINÁMICA (MODIFICADA) =====
                next_step_index = current_step_index
                if next_step_index < len(workflow):
                    next_node_name = workflow[next_step_index]
                    next_view_name = ALL_STEPS[next_node_name]['view_name']
                    # El siguiente paso ('antecedentes_personales_crear') necesita los mismos kwargs
                    next_kwargs = {'historia_id': historia_id, 'tipo': tipo, 'objeto_id': objeto_id}
                    redirect_url = reverse(next_view_name, kwargs=next_kwargs)
                else:
                    redirect_url = reverse('index') # Fallback
                # ===== FIN DE LA LÓGICA DE REDIRECCIÓN DINÁMICA =====


                
                if is_ajax:
                    return JsonResponse({'success': True, 'redirect_url': redirect_url})
                return redirect(redirect_url)
            
            except Exception as e:
                error_msg = f'Error al guardar enfermedad actual: {str(e)}'
                if is_ajax: return JsonResponse({'success': False, 'errors': {'__all__': [error_msg]}}, status=400)
                messages.error(request, error_msg)
                request.session['form_data'] = request.POST.copy()
                return redirect(request.path_info)
        else:
            if is_ajax: return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            messages.error(request, "No se pudo guardar la enfermedad actual. Corrija los errores.")
            context['form'] = form # Añadir el formulario con errores al contexto
            return render(request, 'enfermedad_actual.html', context)

    else: # Lógica GET
        form_data = request.session.pop('form_data', None)
        form = EnfermedadActualForm(form_data or None, instance=antecedentes_instance)
        if editing and not form_data:
            messages.info(request, f"Editando enfermedad actual para {context_object_name}.")

    # Actualizamos el 'form' en el contexto y renderizamos
    context['form'] = form
    return render(request, 'enfermedad_actual.html', context)



# views.py
@require_POST
@login_required
@genetista_or_admin_required
def delete_historia(request, historia_id):
    """
    Vista para eliminar una historia clínica y todos sus datos asociados.
    """
    historia = get_object_or_404(HistoriasClinicas, pk=historia_id)
    user_profile = request.user.genetistas

    # Permiso: Solo el Admin o el genetista dueño de la historia pueden eliminarla.
    if user_profile.rol == 'GEN' and historia.genetista != user_profile:
        messages.error(request, "No tiene permiso para eliminar esta historia clínica.")
        return redirect('ver_historias')

    try:
        numero_historia = historia.numero_historia
        historia.delete()
        messages.success(request, f"La historia clínica N° {numero_historia} y todos sus datos asociados han sido eliminados permanentemente.")
    except Exception as e:
        messages.error(request, f"Ocurrió un error al intentar eliminar la historia: {e}")

    return redirect('ver_historias')

@login_required
@all_roles_required
def buscar_propositos(request):
    query = request.GET.get('q', '').strip()
    propositos_qs = Propositos.objects.none()
    
    try:
        user_gen_profile = request.user.genetistas
    except Genetistas.DoesNotExist:
        return JsonResponse({'propositos': [], 'error': 'Perfil de genetista no encontrado.'}, status=403)

    base_query = Propositos.objects.select_related('historia', 'historia__genetista__user')

    if user_gen_profile.rol == 'ADM' or request.user.is_superuser:
        propositos_qs = base_query.filter(historia__isnull=False)
    elif user_gen_profile.rol == 'GEN':
        propositos_qs = base_query.filter(historia__genetista=user_gen_profile)
    elif user_gen_profile.rol == 'LEC':
        if user_gen_profile.associated_genetista:
            propositos_qs = base_query.filter(historia__genetista=user_gen_profile.associated_genetista)
        else: propositos_qs = Propositos.objects.none()
    else: propositos_qs = Propositos.objects.none()

    if query:
        propositos_qs = propositos_qs.filter(
            Q(nombres__icontains=query) | Q(apellidos__icontains=query) | Q(identificacion__icontains=query)
        )
    
    propositos_qs = propositos_qs.order_by('-historia__fecha_ingreso')[:10 if query else 5]
    resultados = [{
        'proposito_id': p.proposito_id, 'nombres': p.nombres, 'apellidos': p.apellidos,
        'edad': p.edad, 'direccion': p.direccion or "N/A", 'identificacion': p.identificacion,
        'foto_url': p.foto.url if p.foto else None,
        'fecha_ingreso': p.historia.fecha_ingreso.strftime("%d/%m/%Y %H:%M") if p.historia and p.historia.fecha_ingreso else "--",
        'historia_numero': p.historia.numero_historia if p.historia else "N/A",
        'genetista_nombre': p.historia.genetista.user.get_full_name() or p.historia.genetista.user.username if p.historia and p.historia.genetista else "N/A"
    } for p in propositos_qs]
    return JsonResponse({'propositos': resultados})

@never_cache
def signup(request):
    if request.user.is_authenticated: return redirect('index')
    if request.method == 'POST':
        form = ExtendedUserCreationForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    user = form.save()
                login(request, user)
                messages.success(request, "Registro exitoso. ¡Bienvenido!")
                return redirect('index')
            except IntegrityError: messages.error(request, "Nombre de usuario ya existe o error de BD.")
            except Exception as e: messages.error(request, f"Error inesperado: {e}")
        else: messages.error(request, "No se pudo completar registro. Corrija errores.")
    else: form = ExtendedUserCreationForm()
    return render(request, "signup.html", {'form': form})

@never_cache
def login_medico(request):
    if request.user.is_authenticated:
        return redirect('index')
    
    # ===== INICIO DE LA MODIFICACIÓN =====
    error_message = None  # Variable para el error de credenciales
    
    if request.method == 'POST':
        form = LoginForm(request.POST) 
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                try:
                    gen_profile = getattr(user, 'genetistas', None)
                    if not gen_profile:
                        if user.is_superuser:
                           gen_profile, _ = Genetistas.objects.get_or_create(user=user)
                        else:
                           # Este es un error de configuración, usamos messages.error porque es importante
                           messages.error(request, "Perfil de aplicación no encontrado. Contacte al administrador.")
                           return render(request, "login.html", {'form': form})
                    
                    if not gen_profile.rol:
                        # Error de configuración, también usamos messages.error
                        messages.error(request, "El perfil no tiene un rol asignado. Contacte al administrador.")
                        return render(request, "login.html", {'form': form})

                    login(request, user)
                    # Mensaje de bienvenida, este sí va a la siguiente página
                    messages.success(request, f"Bienvenido, {user.get_full_name() or user.username} ({gen_profile.get_rol_display()}).")
                    return redirect(request.GET.get('next') or 'index')
                except Genetistas.DoesNotExist:
                     messages.error(request, "Este usuario no tiene un perfil de genetista. Contacte al administrador.")
            else:
                # El error de credenciales se maneja localmente
                error_message = "Usuario o contraseña incorrectos."
        else:
            # El error de formulario inválido también se maneja localmente
            error_message = "Por favor, ingrese un usuario y contraseña válidos."
    else:
        form = LoginForm()

    # Pasamos el error local (si existe) al contexto de la plantilla
    return render(request, "login.html", {'form': form, 'error': error_message})

def signout(request):
    logout(request)
    messages.success(request, "Sesión cerrada exitosamente.")
    return redirect('login')

# En views.py, reemplaza las vistas 'reports_view' y 'export_report_data' con estas:

# En views.py, reemplaza SOLAMENTE la función 'reports_view' con esta:

@login_required
@all_roles_required
@never_cache
def reports_view(request):
    clear_editing_session(request)
    form = ReportSearchForm(request.GET or None, user=request.user)
    
    results, headers = [], []
    search_attempted = bool(request.GET)
    report_type_display_name = ''

    if form.is_valid():
        report_type = form.cleaned_data.get('report_type', 'patients')
        report_type_display_name = dict(form.fields['report_type'].choices).get(report_type, '')
        
        user_gen_profile = request.user.genetistas
        genetista_obj = form.cleaned_data.get('genetista')
        date_range = form.cleaned_data.get('date_range')
        
        if report_type == 'histories':
            # ... (sin cambios en este bloque) ...
            headers = ['N° Historia', 'Paciente(s)', 'Motivo', 'Estado', 'Genetista', 'Fecha Creación', 'Última Modificación']
            qs = HistoriasClinicas.objects.select_related('genetista__user').prefetch_related('propositos_set')
            
            if user_gen_profile.rol == 'GEN': qs = qs.filter(genetista=user_gen_profile)
            elif user_gen_profile.rol == 'LEC' and user_gen_profile.associated_genetista: qs = qs.filter(genetista=user_gen_profile.associated_genetista)
            elif genetista_obj: qs = qs.filter(genetista=genetista_obj)

            if form.cleaned_data.get('estado_historia'): qs = qs.filter(estado=form.cleaned_data.get('estado_historia'))
            if form.cleaned_data.get('numero_historia_desde'): qs = qs.filter(numero_historia__gte=form.cleaned_data.get('numero_historia_desde'))
            if form.cleaned_data.get('numero_historia_hasta'): qs = qs.filter(numero_historia__lte=form.cleaned_data.get('numero_historia_hasta'))
            if date_range:
                if date_range.get('desde'): qs = qs.filter(fecha_ingreso__date__gte=date_range['desde'])
                if date_range.get('hasta'): qs = qs.filter(fecha_ingreso__date__lte=date_range['hasta'])

            for h in qs.order_by('-numero_historia'):
                results.append([f"HC-{h.numero_historia}", h.get_paciente_display(), h.get_motivo_tipo_consulta_display(), h.get_estado_display(), h.genetista.user.get_full_name() if h.genetista else "N/A", h.fecha_ingreso.strftime('%d-%m-%Y'), h.fecha_ultima_modificacion.strftime('%d-%m-%Y %H:%M') if h.fecha_ultima_modificacion else "N/A"])
        
        else:
            base_propositos_qs = _get_pacientes_queryset_for_role(request.user)
            if genetista_obj: base_propositos_qs = base_propositos_qs.filter(historia__genetista=genetista_obj)

            if report_type == 'patients':
                headers = ['N° Historia', 'Paciente', 'Identificación', 'Edad', 'Genetista', 'Fecha Ingreso', 'Estado']
                if form.cleaned_data.get('buscar_paciente'):
                    q = form.cleaned_data.get('buscar_paciente')
                    base_propositos_qs = base_propositos_qs.filter(Q(nombres__icontains=q) | Q(apellidos__icontains=q) | Q(identificacion__icontains=q))
                if form.cleaned_data.get('estado_paciente'):
                    base_propositos_qs = base_propositos_qs.filter(estado=form.cleaned_data.get('estado_paciente'))
                if date_range:
                    if date_range.get('desde'): base_propositos_qs = base_propositos_qs.filter(historia__fecha_ingreso__date__gte=date_range['desde'])
                    if date_range.get('hasta'): base_propositos_qs = base_propositos_qs.filter(historia__fecha_ingreso__date__lte=date_range['hasta'])
                
                for p in base_propositos_qs.order_by('-historia__fecha_ingreso'):
                    # --- INICIO DE LA MODIFICACIÓN ---
                    edad_display = get_edad_display(p.fecha_nacimiento)
                    results.append([f"HC-{p.historia.numero_historia}", f"{p.nombres} {p.apellidos}", p.identificacion, edad_display, p.historia.genetista.user.get_full_name() if p.historia.genetista else "N/A", p.historia.fecha_ingreso.strftime('%d-%m-%Y'), p.get_estado_display()])
                    # --- FIN DE LA MODIFICACIÓN ---
            
            elif report_type == 'consultations':
                 # ... (sin cambios en este bloque) ...
                headers = ['Paciente', 'N° Historia', 'Plan de Estudio', 'Estado', 'Fecha Próxima Visita', 'Genetista']
                qs = PlanEstudio.objects.filter(Q(evaluacion__proposito__in=base_propositos_qs) | Q(evaluacion__pareja__proposito_id_1__in=base_propositos_qs)).select_related('evaluacion__proposito__historia__genetista__user','evaluacion__pareja__proposito_id_1__historia__genetista__user').distinct()
                if form.cleaned_data.get('estado_consulta') == 'pendientes': qs = qs.filter(completado=False)
                elif form.cleaned_data.get('estado_consulta') == 'completadas': qs = qs.filter(completado=True)
                if date_range:
                    if date_range.get('desde'): qs = qs.filter(fecha_visita__gte=date_range['desde'])
                    if date_range.get('hasta'): qs = qs.filter(fecha_visita__lte=date_range['hasta'])
                for plan in qs.order_by('-fecha_visita'):
                    paciente, hc, genetista = "N/A", "N/A", "N/A"
                    if plan.evaluacion.proposito: paciente, hc, genetista = f"{plan.evaluacion.proposito.nombres} {plan.evaluacion.proposito.apellidos}", f"HC-{plan.evaluacion.proposito.historia.numero_historia}", plan.evaluacion.proposito.historia.genetista.user.get_full_name() if plan.evaluacion.proposito.historia.genetista else "N/A"
                    elif plan.evaluacion.pareja: p1=plan.evaluacion.pareja.proposito_id_1; paciente, hc, genetista = f"Pareja: {p1.nombres} y {plan.evaluacion.pareja.proposito_id_2.nombres}", f"HC-{p1.historia.numero_historia}", p1.historia.genetista.user.get_full_name() if p1.historia.genetista else "N/A"
                    results.append([paciente, hc, plan.accion, "Completada" if plan.completado else "Pendiente", plan.fecha_visita.strftime('%d-%m-%Y') if plan.fecha_visita else "N/A", genetista])
            
            elif report_type == 'diagnoses':
                # ... (sin cambios en este bloque) ...
                headers = ['Paciente', 'N° Historia', 'Diagnóstico Final', 'Genetista', 'Fecha de Evaluación']
                qs = EvaluacionGenetica.objects.filter(Q(proposito__in=base_propositos_qs) | Q(pareja__proposito_id_1__in=base_propositos_qs)).filter(diagnostico_final__isnull=False).exclude(diagnostico_final__exact='').select_related('proposito__historia__genetista__user','pareja__proposito_id_1__historia__genetista__user').distinct()
                if form.cleaned_data.get('buscar_diagnostico'): qs = qs.filter(diagnostico_final__icontains=form.cleaned_data.get('buscar_diagnostico'))
                if date_range:
                    if date_range.get('desde'): qs = qs.filter(fecha_creacion__date__gte=date_range['desde'])
                    if date_range.get('hasta'): qs = qs.filter(fecha_creacion__date__lte=date_range['hasta'])
                for ev in qs.order_by('-fecha_creacion'):
                    paciente, hc, genetista, fecha_ev = "N/A", "N/A", "N/A", "N/A"
                    if ev.proposito: paciente, hc, genetista, fecha_ev = f"{ev.proposito.nombres} {ev.proposito.apellidos}", f"HC-{ev.proposito.historia.numero_historia}", ev.proposito.historia.genetista.user.get_full_name() if ev.proposito.historia.genetista else "N/A", ev.fecha_creacion.strftime('%d-%m-%Y')
                    elif ev.pareja: p1 = ev.pareja.proposito_id_1; paciente, hc, genetista, fecha_ev = f"Pareja: {p1.nombres} y {ev.pareja.proposito_id_2.nombres}", f"HC-{p1.historia.numero_historia}", p1.historia.genetista.user.get_full_name() if p1.historia.genetista else "N/A", ev.fecha_creacion.strftime('%d-%m-%Y')
                    results.append([paciente, hc, ev.diagnostico_final, genetista, fecha_ev])

    context = {
        'form': form, 'results': results, 'headers': headers,
        'search_attempted': search_attempted, 'report_type_display_name': report_type_display_name,
    }
    return render(request, 'reports.html', context)

# --- VISTA 'export_report_data' MODIFICADA ---
# --- VISTA 'export_report_data' MODIFICADA ---
@login_required
@all_roles_required
def export_report_data(request, export_format):
    form = ReportSearchForm(request.GET or None, user=request.user)
    if not form.is_valid():
        return HttpResponse("Filtros inválidos para exportación.", status=400)

    results, headers = [], []
    report_type = form.cleaned_data.get('report_type', 'patients')
    
    user_gen_profile = request.user.genetistas
    genetista_obj = form.cleaned_data.get('genetista')
    date_range = form.cleaned_data.get('date_range')
    
    if report_type == 'histories':
        headers = ['N° Historia', 'Paciente(s)', 'Motivo', 'Estado', 'Genetista', 'Fecha Creación', 'Última Modificación']
        qs = HistoriasClinicas.objects.select_related('genetista__user').prefetch_related('propositos_set')
        
        if user_gen_profile.rol == 'GEN': qs = qs.filter(genetista=user_gen_profile)
        elif user_gen_profile.rol == 'LEC' and user_gen_profile.associated_genetista: qs = qs.filter(genetista=user_gen_profile.associated_genetista)
        elif genetista_obj: qs = qs.filter(genetista=genetista_obj)

        if form.cleaned_data.get('estado_historia'): qs = qs.filter(estado=form.cleaned_data.get('estado_historia'))
        if form.cleaned_data.get('numero_historia_desde'): qs = qs.filter(numero_historia__gte=form.cleaned_data.get('numero_historia_desde'))
        if form.cleaned_data.get('numero_historia_hasta'): qs = qs.filter(numero_historia__lte=form.cleaned_data.get('numero_historia_hasta'))
        if date_range:
            if date_range.get('desde'): qs = qs.filter(fecha_ingreso__date__gte=date_range['desde'])
            if date_range.get('hasta'): qs = qs.filter(fecha_ingreso__date__lte=date_range['hasta'])
        for h in qs.order_by('-numero_historia'):
            results.append([f"HC-{h.numero_historia}", h.get_paciente_display(), h.get_motivo_tipo_consulta_display(), h.get_estado_display(), h.genetista.user.get_full_name() if h.genetista else "N/A", h.fecha_ingreso.strftime('%d-%m-%Y'), h.fecha_ultima_modificacion.strftime('%d-%m-%Y %H:%M') if h.fecha_ultima_modificacion else "N/A"])
    
    else:
        base_propositos_qs = _get_pacientes_queryset_for_role(request.user)
        if genetista_obj: base_propositos_qs = base_propositos_qs.filter(historia__genetista=genetista_obj)

        if report_type == 'patients':
            headers = ['N° Historia', 'Paciente', 'Identificación', 'Edad', 'Genetista', 'Fecha Ingreso', 'Estado']
            if form.cleaned_data.get('buscar_paciente'):
                q = form.cleaned_data.get('buscar_paciente')
                base_propositos_qs = base_propositos_qs.filter(Q(nombres__icontains=q) | Q(apellidos__icontains=q) | Q(identificacion__icontains=q))
            if form.cleaned_data.get('estado_paciente'):
                base_propositos_qs = base_propositos_qs.filter(estado=form.cleaned_data.get('estado_paciente'))
            if date_range:
                if date_range.get('desde'): base_propositos_qs = base_propositos_qs.filter(historia__fecha_ingreso__date__gte=date_range['desde'])
                if date_range.get('hasta'): base_propositos_qs = base_propositos_qs.filter(historia__fecha_ingreso__date__lte=date_range['hasta'])
            
            for p in base_propositos_qs.order_by('-historia__fecha_ingreso'):
                edad_display = get_edad_display(p.fecha_nacimiento)
                results.append([f"HC-{p.historia.numero_historia}", f"{p.nombres} {p.apellidos}", p.identificacion, edad_display, p.historia.genetista.user.get_full_name() if p.historia.genetista else "N/A", p.historia.fecha_ingreso.strftime('%d-%m-%Y'), p.get_estado_display()])
        
        elif report_type == 'consultations':
            headers = ['Paciente', 'N° Historia', 'Plan de Estudio', 'Estado', 'Fecha Próxima Visita', 'Genetista']
            qs = PlanEstudio.objects.filter(Q(evaluacion__proposito__in=base_propositos_qs) | Q(evaluacion__pareja__proposito_id_1__in=base_propositos_qs)).select_related('evaluacion__proposito__historia__genetista__user','evaluacion__pareja__proposito_id_1__historia__genetista__user').distinct()
            if form.cleaned_data.get('estado_consulta') == 'pendientes': qs = qs.filter(completado=False)
            elif form.cleaned_data.get('estado_consulta') == 'completadas': qs = qs.filter(completado=True)
            if date_range:
                if date_range.get('desde'): qs = qs.filter(fecha_visita__gte=date_range['desde'])
                if date_range.get('hasta'): qs = qs.filter(fecha_visita__lte=date_range['hasta'])

            for plan in qs.order_by('-fecha_visita'):
                paciente, hc, genetista = "N/A", "N/A", "N/A"
                if plan.evaluacion.proposito: paciente, hc, genetista = f"{plan.evaluacion.proposito.nombres} {plan.evaluacion.proposito.apellidos}", f"HC-{plan.evaluacion.proposito.historia.numero_historia}", plan.evaluacion.proposito.historia.genetista.user.get_full_name() if plan.evaluacion.proposito.historia.genetista else "N/A"
                elif plan.evaluacion.pareja: p1=plan.evaluacion.pareja.proposito_id_1; paciente, hc, genetista = f"Pareja: {p1.nombres} y {plan.evaluacion.pareja.proposito_id_2.nombres}", f"HC-{p1.historia.numero_historia}", p1.historia.genetista.user.get_full_name() if p1.historia.genetista else "N/A"
                results.append([paciente, hc, plan.accion, "Completada" if plan.completado else "Pendiente", plan.fecha_visita.strftime('%d-%m-%Y') if plan.fecha_visita else "N/A", genetista])
        elif report_type == 'diagnoses':
            headers = ['Paciente', 'N° Historia', 'Diagnóstico Final', 'Genetista', 'Fecha de Evaluación']
            qs = EvaluacionGenetica.objects.filter(Q(proposito__in=base_propositos_qs) | Q(pareja__proposito_id_1__in=base_propositos_qs)).filter(diagnostico_final__isnull=False).exclude(diagnostico_final__exact='').select_related('proposito__historia__genetista__user','pareja__proposito_id_1__historia__genetista__user').distinct()
            if form.cleaned_data.get('buscar_diagnostico'): qs = qs.filter(diagnostico_final__icontains=form.cleaned_data.get('buscar_diagnostico'))
            if date_range:
                if date_range.get('desde'): qs = qs.filter(fecha_creacion__date__gte=date_range['desde'])
                if date_range.get('hasta'): qs = qs.filter(fecha_creacion__date__lte=date_range['hasta'])
            for ev in qs.order_by('-fecha_creacion'):
                paciente, hc, genetista, fecha_ev = "N/A", "N/A", "N/A", "N/A"
                if ev.proposito: paciente, hc, genetista, fecha_ev = f"{ev.proposito.nombres} {ev.proposito.apellidos}", f"HC-{ev.proposito.historia.numero_historia}", ev.proposito.historia.genetista.user.get_full_name() if ev.proposito.historia.genetista else "N/A", ev.fecha_creacion.strftime('%d-%m-%Y')
                elif ev.pareja: p1 = ev.pareja.proposito_id_1; paciente, hc, genetista, fecha_ev = f"Pareja: {p1.nombres} y {ev.pareja.proposito_id_2.nombres}", f"HC-{p1.historia.numero_historia}", p1.historia.genetista.user.get_full_name() if p1.historia.genetista else "N/A", ev.fecha_creacion.strftime('%d-%m-%Y')
                results.append([paciente, hc, ev.diagnostico_final, genetista, fecha_ev])
    
    # --- Lógica de exportación MODIFICADA ---
    if export_format == 'csv': # El endpoint sigue siendo 'csv' pero ahora genera un archivo Excel (.xlsx)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte_{report_type}"

        # Estilos para la cabecera
        header_font = Font(name='Arial', bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Escribir cabeceras y aplicar estilo
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        # Escribir filas de datos
        for row_data in results:
            ws.append(row_data)
        
        # Ajustar el ancho de las columnas después de haber insertado todos los datos
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Crear la respuesta HTTP con el archivo XLSX
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response) # Guardar el libro de trabajo directamente en la respuesta
        return response

    elif export_format == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements, styles = [], getSampleStyleSheet()
        elements.append(Paragraph(f" {dict(form.fields['report_type'].choices).get(report_type)}", styles['h1']))
        elements.append(Spacer(1, 0.2*72))
        if not headers:
            headers = ["Error"]; results = [["Tipo de reporte no válido o sin datos."]]
        table_data = [headers] + results
        if not results:
             table_data.append(["No se encontraron registros."] * len(headers))
        table = Table(table_data); table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black), ('FONTSIZE', (0,0), (-1,-1), 8), ])); elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
        
    return HttpResponse("Formato de exportación no soportado.", status=400)

@login_required
@admin_required
def toggle_user_active_status(request, user_id):
    if request.method == 'POST':
        user_to_toggle = get_object_or_404(User, pk=user_id)
        if user_to_toggle == request.user:
            messages.error(request, "No puede cambiar el estado de su propia cuenta.")
        else:
            user_to_toggle.is_active = not user_to_toggle.is_active
            user_to_toggle.save()
            status_message = "activado" if user_to_toggle.is_active else "desactivado"
            messages.success(request, f"Usuario {user_to_toggle.username} {status_message} exitosamente.")
    return redirect('gestion_usuarios')

@login_required
@admin_required
def delete_user_admin(request, user_id):
    if request.method == 'POST':
        user_to_delete = get_object_or_404(User, pk=user_id)
        if user_to_delete == request.user:
            messages.error(request, "No puede eliminar su propia cuenta.")
        else:
            try:
                with transaction.atomic():
                    username = user_to_delete.username
                    user_to_delete.delete()
                    messages.success(request, f"Usuario {username} eliminado exitosamente.")
            except Exception as e:
                messages.error(request, f"Error al eliminar usuario: {e}")
    return redirect('gestion_usuarios')

def _get_pacientes_queryset_for_role(user):
    """
    Función auxiliar para obtener el queryset de pacientes basado en el rol del usuario.
    Centraliza la lógica de permisos de acceso a datos.
    """
    try:
        user_gen_profile = user.genetistas
    except Genetistas.DoesNotExist:
        # Si es superusuario y no tiene perfil, se lo creamos con rol ADM
        if user.is_superuser:
            user_gen_profile, _ = Genetistas.objects.get_or_create(user=user, defaults={'rol': 'ADM'})
        else:
            return Propositos.objects.none()

    role = user_gen_profile.rol
    base_qs = Propositos.objects.filter(historia__isnull=False).select_related('historia', 'historia__genetista__user')

    if role == 'ADM' or user.is_superuser:
        return base_qs
    elif role == 'GEN':
        return base_qs.filter(historia__genetista=user_gen_profile)
    elif role == 'LEC':
        if user_gen_profile.associated_genetista:
            return base_qs.filter(historia__genetista=user_gen_profile.associated_genetista)
        else:
            # Lector no asociado no ve nada
            return Propositos.objects.none()
    
    return Propositos.objects.none()



@require_POST
@login_required
@admin_required
def reset_password_admin(request, user_id):
    user_to_reset = get_object_or_404(User, pk=user_id)
    form = PasswordResetAdminForm(request.POST)

    if form.is_valid():
        new_password = form.cleaned_data['new_password']
        
        # 1. Guardar la nueva contraseña en la base de datos
        user_to_reset.set_password(new_password)
        user_to_reset.save()

        # 2. Intentar enviar el correo de notificación
        if user_to_reset.email:
            try:
                # Preparamos el contexto para las plantillas de correo
                context = {
                    'user': user_to_reset,
                    'new_password': new_password
                }
                
                # Renderizamos tanto la plantilla HTML como la de texto plano
                html_message = render_to_string('emails/password_reset_notification.html', context)
                plain_message = render_to_string('emails/password_reset_notification.txt', context)

                send_mail(
                    subject='Tu contraseña ha sido restablecida - IIG LUZ',
                    message=plain_message,  # Mensaje de texto plano como fallback
                    from_email=settings.EMAIL_HOST_USER, # Usamos el correo configurado en settings
                    recipient_list=[user_to_reset.email],
                    html_message=html_message, # Adjuntamos la versión HTML
                    fail_silently=False, # Si falla, lanzará una excepción
                )
                messages.success(request, f"Contraseña para '{user_to_reset.username}' restablecida y notificación enviada a {user_to_reset.email}.")

            except Exception as e:
                # Si el envío falla, la contraseña ya fue cambiada, pero informamos al admin del error.
                messages.warning(request, f"La contraseña para '{user_to_reset.username}' fue restablecida, pero falló el envío del correo de notificación. Revisa la configuración del servidor. Error: {e}")
        
        else:
            # Si el usuario no tiene email, solo informamos que no se pudo notificar.
            messages.info(request, f"La contraseña para '{user_to_reset.username}' fue restablecida. No se envió notificación porque el usuario no tiene un correo registrado.")

    else:
        # Si el formulario no es válido (ej. contraseñas no coinciden), mostramos un error más específico.
        error_message = form.errors.get('__all__') or ["Error desconocido en el formulario."]
        messages.error(request, f"Error al restablecer la contraseña: {error_message[0]}")

    return redirect('gestion_usuarios')


@login_required
@all_roles_required
def index_view(request):
    clear_editing_session(request)
    user_gen_profile = request.user.genetistas
    role = user_gen_profile.rol
    
    stats = {'historias': 0, 'pacientes': 0, 'consultas_completadas': 0, 'diagnosticos_completados': 0}
    
    target_genetista = None
    if role == 'GEN':
        target_genetista = user_gen_profile
    elif role == 'LEC':
        target_genetista = user_gen_profile.associated_genetista
    
    if role == 'ADM' or request.user.is_superuser:
        stats['historias'] = HistoriasClinicas.objects.count()
        stats['pacientes'] = Propositos.objects.count()
        stats['consultas_completadas'] = PlanEstudio.objects.filter(completado=True).count()
        stats['diagnosticos_completados'] = EvaluacionGenetica.objects.filter(diagnostico_final__isnull=False).exclude(diagnostico_final__exact='').count()

    elif target_genetista:
        stats['historias'] = HistoriasClinicas.objects.filter(genetista=target_genetista).count()
        stats['pacientes'] = Propositos.objects.filter(historia__genetista=target_genetista).count()
        
        q_proposito_plan = Q(evaluacion__proposito__historia__genetista=target_genetista)
        q_pareja_plan = Q(evaluacion__pareja__proposito_id_1__historia__genetista=target_genetista) | Q(evaluacion__pareja__proposito_id_2__historia__genetista=target_genetista)
        stats['consultas_completadas'] = PlanEstudio.objects.filter(completado=True).filter(q_proposito_plan | q_pareja_plan).distinct().count()

        q_proposito_diag = Q(proposito__historia__genetista=target_genetista)
        q_pareja_diag = Q(pareja__proposito_id_1__historia__genetista=target_genetista) | Q(pareja__proposito_id_2__historia__genetista=target_genetista)
        stats['diagnosticos_completados'] = EvaluacionGenetica.objects.filter(
            diagnostico_final__isnull=False
        ).exclude(diagnostico_final__exact='').filter(
            q_proposito_diag | q_pareja_diag
        ).distinct().count()

    recent_activities = []
    historias_qs = HistoriasClinicas.objects.select_related('genetista__user')
    planes_completados_qs = PlanEstudio.objects.filter(completado=True).select_related('evaluacion__proposito', 'evaluacion__pareja__proposito_id_1', 'evaluacion__pareja__proposito_id_2')

    last_historias, last_planes_completados = HistoriasClinicas.objects.none(), PlanEstudio.objects.none()

    if role == 'ADM' or request.user.is_superuser:
        last_historias = historias_qs.order_by('-fecha_ingreso')[:5]
        last_planes_completados = planes_completados_qs.order_by('-fecha_visita')[:5]
    elif target_genetista:
        last_historias = historias_qs.filter(genetista=target_genetista).order_by('-fecha_ingreso')[:5]
        q_proposito_plan_recent = Q(evaluacion__proposito__historia__genetista=target_genetista)
        q_pareja_plan_recent = Q(evaluacion__pareja__proposito_id_1__historia__genetista=target_genetista) | Q(evaluacion__pareja__proposito_id_2__historia__genetista=target_genetista)
        last_planes_completados = planes_completados_qs.filter(q_proposito_plan_recent | q_pareja_plan_recent).distinct().order_by('-fecha_visita')[:5]

    for historia in last_historias:
        proposito = Propositos.objects.filter(historia=historia).first()
        paciente_name = f"{proposito.nombres} {proposito.apellidos}" if proposito else f"Historia N° {historia.numero_historia}"
        recent_activities.append({'title': 'Nueva historia clínica creada', 'meta': f'Paciente: {paciente_name}', 'timestamp': historia.fecha_ingreso, 'indicator_class': 'success'})
    
    for plan in last_planes_completados:
        paciente_name = "N/A"
        if plan.evaluacion.proposito:
            paciente_name = f"{plan.evaluacion.proposito.nombres} {plan.evaluacion.proposito.apellidos}"
        elif plan.evaluacion.pareja:
            p1, p2 = plan.evaluacion.pareja.proposito_id_1, plan.evaluacion.pareja.proposito_id_2
            paciente_name = f"Pareja: {p1.nombres} y {p2.nombres}"
        
        if plan.fecha_visita:
            naive_datetime = datetime.combine(plan.fecha_visita, time.min)
            aware_datetime = timezone.make_aware(naive_datetime)
            recent_activities.append({'title': 'Análisis genético completado', 'meta': f'Paciente: {paciente_name}', 'timestamp': aware_datetime, 'indicator_class': 'info'})
    
    recent_activities = sorted([act for act in recent_activities if act['timestamp']], key=lambda x: x['timestamp'], reverse=True)[:5]
    
    page_title = f"Inicio ({user_gen_profile.get_rol_display()})" if user_gen_profile else "Inicio"
    if role == 'LEC' and not target_genetista:
        messages.warning(request, "Usted es un Lector no asociado a ningún Genetista. Su vista de datos estará limitada.")

    context = {'page_title': page_title, 'can_create_historia': role in ['GEN', 'ADM'], 'stats': stats, 'recent_activities': recent_activities}
    return render(request, "index.html", context)


@login_required
@admin_required
def pacientes_admin_view(request):
    ultimos_propositos_qs = _get_pacientes_queryset_for_role(request.user)
    context = {'ultimos_propositos': ultimos_propositos_qs[:10], 'page_title': "Lista de Pacientes (Administrador)", 'can_create_historia': True}
    return render(request, "index.html", context)

@login_required
@genetista_required
def pacientes_genetista_view(request):
    ultimos_propositos_qs = _get_pacientes_queryset_for_role(request.user)
    context = {'ultimos_propositos': ultimos_propositos_qs[:5], 'page_title': "Mis Pacientes", 'can_create_historia': True}
    return render(request, "index.html", context)

@login_required
@lector_required
def pacientes_lector_view(request, genetista_id):
    user_profile = request.user.genetistas
    target_genetista = get_object_or_404(Genetistas, pk=genetista_id, rol='GEN')

    if user_profile.rol == 'LEC' and (not user_profile.associated_genetista or user_profile.associated_genetista.pk != genetista_id):
        raise PermissionDenied("No tiene permiso para ver pacientes de este genetista.")

    ultimos_propositos_qs = Propositos.objects.filter(historia__genetista=target_genetista).select_related('historia').order_by('-historia__fecha_ingreso')
    context = {'ultimos_propositos': ultimos_propositos_qs[:5], 'page_title': f"Pacientes de {target_genetista.user.get_full_name()}", 'can_create_historia': False}
    return render(request, "index.html", context)

@login_required
@all_roles_required
def pacientes_redirect_view(request):
    clear_editing_session(request)
    user_profile = request.user.genetistas
    role = user_profile.rol

    if role == 'ADM': return redirect('pacientes_admin_list')
    elif role == 'GEN': return redirect('pacientes_genetista_list')
    elif role == 'LEC':
        if user_profile.associated_genetista_id:
            return redirect('pacientes_lector_list', genetista_id=user_profile.associated_genetista_id)
        else:
            messages.warning(request, "Como Lector, no está asociado a ningún Genetista.")
            return redirect('index')
    else: return redirect('index')
    
# En views.py

@login_required
@all_roles_required
@never_cache
def gestion_pacientes_view(request):
    base_pacientes_qs = _get_pacientes_queryset_for_role(request.user)

    # Estadísticas (sin cambios)
    total_pacientes = base_pacientes_qs.count()
    pacientes_cerrados = base_pacientes_qs.filter(estado=Propositos.ESTADO_CERRADO).count()
    pacientes_seguimiento = base_pacientes_qs.filter(estado=Propositos.ESTADO_SEGUIMIENTO).count()
    pacientes_inactivos = base_pacientes_qs.filter(estado=Propositos.ESTADO_INACTIVO).count()
    
    historias_ids = base_pacientes_qs.values_list('historia_id', flat=True).distinct()
    historias_clinicas_count = HistoriasClinicas.objects.filter(pk__in=historias_ids).count()
    evaluaciones_ids = EvaluacionGenetica.objects.filter(Q(proposito__in=base_pacientes_qs) | Q(pareja__proposito_id_1__in=base_pacientes_qs) | Q(pareja__proposito_id_2__in=base_pacientes_qs)).values_list('evaluacion_id', flat=True).distinct()
    analisis_pendientes_count = PlanEstudio.objects.filter(evaluacion_id__in=evaluaciones_ids, completado=False).count()
    now = timezone.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    nuevos_pacientes_mes = base_pacientes_qs.filter(historia__fecha_ingreso__gte=start_of_month).count()
    nuevas_historias_mes = HistoriasClinicas.objects.filter(pk__in=historias_ids, fecha_ingreso__gte=start_of_month).count()
    porcentaje_cerrados = (pacientes_cerrados / total_pacientes * 100) if total_pacientes > 0 else 0

    # Formulario de búsqueda (sin cambios)
    user_role = request.user.genetistas.rol
    estado_choices_for_filter = Propositos.ESTADO_CHOICES if user_role == 'ADM' or request.user.is_superuser else [c for c in Propositos.ESTADO_CHOICES if c[0] != Propositos.ESTADO_INACTIVO]
    form = PatientSearchForm(request.GET or None, user=request.user, estado_choices=estado_choices_for_filter)
    
    pacientes_a_procesar_qs = base_pacientes_qs
    search_attempted = bool(request.GET)
    if form.is_valid():
        if form.cleaned_data.get('genetista'):
            pacientes_a_procesar_qs = pacientes_a_procesar_qs.filter(historia__genetista=form.cleaned_data.get('genetista'))
        if form.cleaned_data.get('buscar_paciente'):
            query = form.cleaned_data.get('buscar_paciente')
            pacientes_a_procesar_qs = pacientes_a_procesar_qs.filter(Q(nombres__icontains=query) | Q(apellidos__icontains=query) | Q(identificacion__icontains=query))
        if form.cleaned_data.get('estado'):
            pacientes_a_procesar_qs = pacientes_a_procesar_qs.filter(estado=form.cleaned_data.get('estado'))

    # ===== INICIO DE LA LÓGICA MODIFICADA PARA AGRUPAR PAREJAS =====
    
    # Anotamos el conteo de planes pendientes para historias de un solo propósito.
    pacientes_con_conteo_single = pacientes_a_procesar_qs.annotate(
        plan_estudio_pendiente_count_single=Count('evaluaciongenetica__planes_estudio', filter=Q(evaluaciongenetica__planes_estudio__completado=False))
    )

    pacientes_dict = {p.proposito_id: p for p in pacientes_con_conteo_single}
    processed_ids = set()
    final_list = []

    parejas_qs = Parejas.objects.filter(
        Q(proposito_id_1_id__in=pacientes_dict.keys()) | Q(proposito_id_2_id__in=pacientes_dict.keys())
    ).select_related(
        'proposito_id_1__historia__genetista__user', 
        'proposito_id_2__historia__genetista__user'
    ).prefetch_related(
        'evaluaciongenetica_set__planes_estudio'
    )

    # Procesamos las parejas primero
    for pareja in parejas_qs:
        if pareja.proposito_id_1_id in pacientes_dict and pareja.proposito_id_2_id in pacientes_dict:
            if pareja.proposito_id_1_id not in processed_ids and pareja.proposito_id_2_id not in processed_ids:
                
                evaluacion = pareja.evaluaciongenetica_set.first()
                pareja.plan_estudio_pendiente_count = sum(1 for plan in evaluacion.planes_estudio.all() if not plan.completado) if evaluacion else 0
                
                pareja.proposito_id_1 = pacientes_dict[pareja.proposito_id_1_id]
                pareja.proposito_id_2 = pacientes_dict[pareja.proposito_id_2_id]

                # --- INICIO DE LA MODIFICACIÓN ---
                pareja.proposito_id_1.edad_display = get_edad_display(pareja.proposito_id_1.fecha_nacimiento)
                pareja.proposito_id_2.edad_display = get_edad_display(pareja.proposito_id_2.fecha_nacimiento)
                # --- FIN DE LA MODIFICACIÓN ---

                final_list.append(pareja)
                processed_ids.add(pareja.proposito_id_1_id)
                processed_ids.add(pareja.proposito_id_2_id)

    # Procesamos los pacientes individuales restantes
    for proposito_id, proposito in pacientes_dict.items():
        if proposito_id not in processed_ids:
            # --- INICIO DE LA MODIFICACIÓN ---
            proposito.edad_display = get_edad_display(proposito.fecha_nacimiento)
            # --- FIN DE LA MODIFICACIÓN ---
            proposito.plan_estudio_pendiente_count = proposito.plan_estudio_pendiente_count_single
            final_list.append(proposito)

    # Reordenamos la lista final por fecha para mantener la consistencia
    final_list.sort(key=lambda item: (item.proposito_id_1.historia.fecha_ingreso if hasattr(item, 'pareja_id') else item.historia.fecha_ingreso), reverse=True)
    
    pacientes_count_display = len(final_list) # El contador ahora refleja "entradas" en la tabla

    # ===== FIN DE LA LÓGICA MODIFICADA =====

    context = {
        'total_pacientes': total_pacientes, 'pacientes_cerrados': pacientes_cerrados, 
        'pacientes_seguimiento': pacientes_seguimiento, 'pacientes_inactivos': pacientes_inactivos,
        'historias_clinicas_count': historias_clinicas_count, 'analisis_pendientes_count': analisis_pendientes_count,
        'nuevos_pacientes_mes': nuevos_pacientes_mes, 'nuevas_historias_mes': nuevas_historias_mes, 
        'porcentaje_cerrados': porcentaje_cerrados,
        'form': form, 
        'pacientes_list': final_list, # Usamos la nueva lista procesada
        'pacientes_count': pacientes_count_display, # Usamos el nuevo contador
        'search_attempted': search_attempted
    }
    return render(request, "gestion_pacientes.html", context)

# ===== INICIO DE CÓDIGO A AÑADIR EN views.py =====
@require_POST
@login_required
@genetista_or_admin_required
def inactivar_pacientes_view(request):
    """
    Vista para inactivar en masa a los pacientes seleccionados.
    """
    try:
        data = json.loads(request.body)
        paciente_ids = data.get('paciente_ids', [])
        
        if not isinstance(paciente_ids, list):
            return JsonResponse({'success': False, 'message': 'Formato de datos inválido.'}, status=400)

        # Asegura que el usuario tenga permiso para modificar estos pacientes.
        allowed_pacientes_qs = _get_pacientes_queryset_for_role(request.user)
        
        # Filtra los IDs enviados para solo incluir aquellos a los que el usuario tiene acceso.
        pacientes_a_inactivar = allowed_pacientes_qs.filter(pk__in=paciente_ids)
        
        count = pacientes_a_inactivar.count()
        
        if count == 0 and len(paciente_ids) > 0:
            return JsonResponse({'success': False, 'message': 'No tiene permiso para modificar los pacientes seleccionados.'}, status=403)

        pacientes_a_inactivar.update(estado=Propositos.ESTADO_INACTIVO)

        message = f'{count} paciente{"s" if count != 1 else ""} {"han" if count != 1 else "ha"} sido inactivado exitosamente.'
        messages.success(request, message) # Este mensaje se mostrará al recargar la página
        return JsonResponse({'success': True, 'message': message})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Solicitud malformada.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Ocurrió un error: {str(e)}'}, status=500)
    
# ===== FIN DEL CÓDIGO A AÑADIR =====
# En views.py

# (Asegúrate de tener todos los imports necesarios al principio del archivo)
# from io import BytesIO
# from reportlab.platypus import ...
# ... etc.

@login_required
@all_roles_required
@never_cache
def generar_pdf_historia(request, historia_id):
    historia = get_object_or_404(HistoriasClinicas, pk=historia_id)
    try:
        user_gen_profile = request.user.genetistas
        if user_gen_profile.rol == 'GEN' and historia.genetista != user_gen_profile: raise PermissionDenied("No tiene permiso para generar el reporte de esta historia.")
        if user_gen_profile.rol == 'LEC':
            if not user_gen_profile.associated_genetista or historia.genetista != user_gen_profile.associated_genetista: raise PermissionDenied("Como lector, no tiene permiso para generar el reporte de esta historia.")
    except Genetistas.DoesNotExist: raise PermissionDenied("Perfil de usuario no encontrado.")

    propositos = list(Propositos.objects.filter(historia=historia).order_by('pk'))
    pareja = Parejas.objects.filter(proposito_id_1__in=propositos, proposito_id_2__in=propositos).first() if len(propositos) > 1 else None
    sujeto_principal = pareja if pareja else (propositos[0] if propositos else None)
    tipo_sujeto = 'pareja' if pareja else 'proposito'

    padres_info = {p.tipo: p for p in InformacionPadres.objects.filter(proposito=sujeto_principal)} if tipo_sujeto == 'proposito' and sujeto_principal else {}
    q_filter = Q(proposito__in=propositos) if propositos else Q(pk__isnull=True)
    if pareja: q_filter |= Q(pareja=pareja)
    
    antecedentes_personales = AntecedentesPersonales.objects.filter(q_filter).first()
    antecedentes_familiares = AntecedentesFamiliaresPreconcepcionales.objects.filter(q_filter).first()
    periodo_neonatal = PeriodoNeonatal.objects.filter(q_filter).first()
    desarrollo_psicomotor = DesarrolloPsicomotor.objects.filter(q_filter).first()
    evaluacion_genetica = EvaluacionGenetica.objects.filter(q_filter).first()
    diagnosticos, planes_estudio = (DiagnosticoPresuntivo.objects.filter(evaluacion=evaluacion_genetica).order_by('orden'), PlanEstudio.objects.filter(evaluacion=evaluacion_genetica).order_by('fecha_visita')) if evaluacion_genetica else (DiagnosticoPresuntivo.objects.none(), PlanEstudio.objects.none())
    examenes_fisicos = ExamenFisico.objects.filter(proposito__in=propositos).select_related('proposito')
    autorizaciones = Autorizaciones.objects.filter(proposito__in=propositos).select_related('proposito', 'representante_padre')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch/2, bottomMargin=inch/2)
    story = []
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Justify', alignment=TA_JUSTIFY)); styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER)); styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT)); styles.add(ParagraphStyle(name='Left', alignment=TA_LEFT)); styles.add(ParagraphStyle(name='SectionTitle', fontSize=12, fontName='Helvetica-Bold', spaceAfter=6)); styles.add(ParagraphStyle(name='SubSectionTitle', fontSize=10, fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=4)); styles.add(ParagraphStyle(name='Alert', fontSize=10, fontName='Helvetica-Bold', textColor=colors.red))

    def _format_val(value, default="N/A"):
        if value is None or value == '': return default
        if isinstance(value, bool): return "Sí" if value else "No"
        if hasattr(value, 'strftime'): return value.strftime('%d/%m/%Y')
        return str(value)
    def _add_title(text): story.append(Paragraph(text, styles['h1'])); story.append(Spacer(1, 0.2*inch))
    def _add_section_title(text): story.append(Spacer(1, 0.2*inch)); story.append(Paragraph(text, styles['SectionTitle'])); story.append(Table([['']], colWidths=[7.5*inch], style=TableStyle([('LINEABOVE', (0,0), (-1,0), 1, colors.black)]))); story.append(Spacer(1, 0.1*inch))
    def _add_subsection_title(text): story.append(Paragraph(text, styles['SubSectionTitle']))
    def _add_key_value_table(data_dict, col_widths=[2.5*inch, 5*inch]):
        table_data = [[Paragraph(f"<b>{k}:</b>", styles['Normal']), Paragraph(_format_val(v), styles['Normal'])] for k, v in data_dict.items()]
        if not table_data: return
        table = Table(table_data, colWidths=col_widths); table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey), ('LEFTPADDING', (0,0), (-1,-1), 6), ('RIGHTPADDING', (0,0), (-1,-1), 6), ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6)])); story.append(table); story.append(Spacer(1, 0.1*inch))

    # ===== INICIO DE LA MODIFICACIÓN 1: Encabezado Genérico =====
    story.append(Spacer(1, 0.3*inch))
    _add_title("HISTORIA CLÍNICA GENÉTICA")
    # ===== FIN DE LA MODIFICACIÓN 1 =====

    _add_key_value_table({"Nº de Historia IIGLUZ": f"HC-{historia.numero_historia}", "Fecha de Ingreso": _format_val(historia.fecha_ingreso), "Tipo de Historia": historia.get_motivo_tipo_consulta_display(), "Genetista": _format_val(historia.genetista.user.get_full_name() if historia.genetista and historia.genetista.user else "No asignado"), "Cursante de Postgrado": _format_val(historia.cursante_postgrado), "Médico Referente": f"{_format_val(historia.medico)} ({_format_val(historia.especialidad)})", "Centro de Referencia": _format_val(historia.centro_referencia),})
    _add_section_title("DATOS DEL PACIENTE(S)")
    if not sujeto_principal: story.append(Paragraph("No hay paciente(s) asignado(s) a esta historia clínica.", styles['Alert']))
    elif tipo_sujeto == 'proposito':
        p = sujeto_principal
        proposito_data = { "Nombres y Apellidos": f"{p.nombres} {p.apellidos}", "Identificación": p.identificacion, "Edad": get_edad_display(p.fecha_nacimiento), "Fecha de Nacimiento": _format_val(p.fecha_nacimiento), "Lugar de Nacimiento": _format_val(p.lugar_nacimiento), "Sexo": _format_val(p.get_sexo_display()), "Escolaridad": _format_val(p.escolaridad), "Ocupación": _format_val(p.ocupacion), "Dirección": _format_val(p.direccion), "Teléfono": _format_val(p.telefono), "Email": _format_val(p.email), "Grupo Sanguíneo": f"{_format_val(p.grupo_sanguineo)} {_format_val(p.factor_rh)}",}
        _add_key_value_table(proposito_data)
        _add_subsection_title("DATOS DEL PADRE")
        padre = padres_info.get('Padre')
        if padre: _add_key_value_table({"Nombres y Apellidos": f"{padre.nombres} {padre.apellidos}", "Identificación": _format_val(padre.identificacion), "Lugar de Nacimiento": _format_val(padre.lugar_nacimiento), "Fecha de Nacimiento": _format_val(padre.fecha_nacimiento), "Grupo Sanguíneo y Factor RH": f"{_format_val(padre.grupo_sanguineo)} {_format_val(padre.factor_rh)}", "Teléfono": _format_val(padre.telefono), "Dirección": _format_val(padre.direccion), "Escolaridad": _format_val(padre.escolaridad), "Ocupación": _format_val(padre.ocupacion),})
        else: story.append(Paragraph("Información del Padre no registrada.", styles['Normal']))
        _add_subsection_title("DATOS DE LA MADRE")
        madre = padres_info.get('Madre')
        if madre: _add_key_value_table({ "Nombres y Apellidos": f"{madre.nombres} {madre.apellidos}", "Identificación": _format_val(madre.identificacion), "Lugar de Nacimiento": _format_val(madre.lugar_nacimiento), "Fecha de Nacimiento": _format_val(madre.fecha_nacimiento), "Grupo Sanguíneo y Factor RH": f"{_format_val(madre.grupo_sanguineo)} {_format_val(madre.factor_rh)}", "Teléfono": _format_val(madre.telefono), "Dirección": _format_val(madre.direccion), "Escolaridad": _format_val(madre.escolaridad), "Ocupación": _format_val(madre.ocupacion),})
        else: story.append(Paragraph("Información de la Madre no registrada.", styles['Normal']))
    elif tipo_sujeto == 'pareja':
        for i, p in enumerate(propositos):
            _add_subsection_title(f"CÓNYUGE {i+1}")
            proposito_data = {"Nombres y Apellidos": f"{p.nombres} {p.apellidos}", "Identificación": p.identificacion, "Edad": get_edad_display(p.fecha_nacimiento), "Fecha de Nacimiento": _format_val(p.fecha_nacimiento), "Escolaridad": _format_val(p.escolaridad), "Ocupación": _format_val(p.ocupacion),}
            _add_key_value_table(proposito_data)
    
    if antecedentes_personales or periodo_neonatal or desarrollo_psicomotor or antecedentes_familiares:
        story.append(PageBreak()); _add_section_title("ANTECEDENTES")
        if antecedentes_personales:
            _add_subsection_title("ANTECEDENTES PRENATALES Y OBSTÉTRICOS")
            ap_data = { "FUR": _format_val(antecedentes_personales.fur), "Edad Gestacional": f"{_format_val(antecedentes_personales.edad_gestacional)} semanas", "Controles Prenatales": _format_val(antecedentes_personales.controles_prenatales), "Nº Gestas/Partos/Cesáreas/Abortos": f"{_format_val(antecedentes_personales.numero_gestas)} / {_format_val(antecedentes_personales.numero_partos)} / {_format_val(antecedentes_personales.numero_cesareas)} / {_format_val(antecedentes_personales.numero_abortos)}", "Complicaciones en Embarazo": _format_val(antecedentes_personales.complicaciones_embarazo), "Exposición a Teratógenos": f"{_format_val(antecedentes_personales.exposicion_teratogenos)}: {_format_val(antecedentes_personales.descripcion_exposicion)}", "Complicaciones en Parto": _format_val(antecedentes_personales.complicaciones_parto), }; _add_key_value_table(ap_data, [2.5*inch, 5*inch])
        if periodo_neonatal:
            _add_subsection_title("PERIODO NEONATAL")
            pn_data = { "Peso al Nacer": f"{_format_val(periodo_neonatal.peso_nacer)} kg", "Talla al Nacer": f"{_format_val(periodo_neonatal.talla_nacer)} cm", "Circ. Cefálica": f"{_format_val(periodo_neonatal.circunferencia_cefalica)} cm", "Complicaciones (Cianosis, Ictericia, etc.)": _format_val(periodo_neonatal.observacion_complicaciones), "Tipo de Alimentación": _format_val(periodo_neonatal.tipo_alimentacion), }; _add_key_value_table(pn_data)
        if desarrollo_psicomotor:
            _add_subsection_title("DESARROLLO PSICOMOTOR")
            dp_data = { "Sostén Cefálico / Sonrisa Social": f"{_format_val(desarrollo_psicomotor.sostener_cabeza)} / {_format_val(desarrollo_psicomotor.sonrisa_social)}", "Sedestación / Gateo": f"{_format_val(desarrollo_psicomotor.sentarse)} / {_format_val(desarrollo_psicomotor.gatear)}", "Marcha / Primeras Palabras": f"{_format_val(desarrollo_psicomotor.caminar)} / {_format_val(desarrollo_psicomotor.primeras_palabras)}", "Progreso Escolar": _format_val(desarrollo_psicomotor.progreso_escuela) }; _add_key_value_table(dp_data)
        if antecedentes_familiares:
            _add_subsection_title("ANTECEDENTES FAMILIARES Y PRECONCEPCIONALES")
            af_data = { "Antecedentes Familiares Paternos": _format_val(antecedentes_familiares.antecedentes_padre), "Antecedentes Familiares Maternos": _format_val(antecedentes_familiares.antecedentes_madre), "Consanguinidad": f"{_format_val(antecedentes_familiares.consanguinidad)} (Grado: {_format_val(antecedentes_familiares.grado_consanguinidad)})", }; _add_key_value_table(af_data)
    if examenes_fisicos:
        story.append(PageBreak())
        for ef in examenes_fisicos:
            _add_section_title(f"EXAMEN FÍSICO - {ef.proposito.nombres} {ef.proposito.apellidos}")
            medidas_data = [ [Paragraph("<b>Medida</b>", styles['Normal']), Paragraph("<b>Valor</b>", styles['Normal']), Paragraph("<b>Medida</b>", styles['Normal']), Paragraph("<b>Valor</b>", styles['Normal'])], ["Peso", f"{_format_val(ef.peso)} kg", "Talla", f"{_format_val(ef.talla)} cm"], ["Circ. Cefálica", f"{_format_val(ef.circunferencia_cefalica)} cm", "Brazada", f"{_format_val(ef.medida_abrazada)} cm"], ["Seg. Superior", f"{_format_val(ef.segmento_superior)} cm", "Seg. Inferior", f"{_format_val(ef.segmento_inferior)} cm"], ["T/A", f"{_format_val(ef.tension_arterial_sistolica)}/{_format_val(ef.tension_arterial_diastolica)} mmHg", "", ""], ]
            t_medidas = Table(medidas_data, colWidths=[1.5*inch, 2.25*inch, 1.5*inch, 2.25*inch]); t_medidas.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 1, colors.black), ('BACKGROUND', (0,0), (-1,0), colors.lightgrey)])); story.append(t_medidas); story.append(Spacer(1, 0.2*inch))
            observaciones_data = { "Cabeza": _format_val(ef.observaciones_cabeza), "Cuello": _format_val(ef.observaciones_cuello), "Tórax": _format_val(ef.observaciones_torax), "Abdomen": _format_val(ef.observaciones_abdomen), "Genitales": _format_val(ef.observaciones_genitales), "Espalda": _format_val(ef.observaciones_espalda), "Miembros": f"Sup: {_format_val(ef.observaciones_miembros_superiores)}<br/>Inf: {_format_val(ef.observaciones_miembros_inferiores)}", "Piel": _format_val(ef.observaciones_piel), "Neurológico": _format_val(ef.observaciones_neurologico), }; _add_key_value_table(observaciones_data, col_widths=[1.5*inch, 6*inch])
    if evaluacion_genetica:
        story.append(PageBreak()); _add_section_title("RESUMEN DE EVALUACIÓN GENÉTICA"); _add_subsection_title("A. SIGNOS CLÍNICOS IMPORTANTES"); story.append(Paragraph(_format_val(evaluacion_genetica.signos_clinicos), styles['Justify'])); _add_subsection_title("B. DIAGNÓSTICOS PRESUNTIVOS")
        if diagnosticos: [story.append(Paragraph(f"{i+1}. {_format_val(diag.descripcion)}", styles['Normal'])) for i, diag in enumerate(diagnosticos)]
        else: story.append(Paragraph("No se registraron diagnósticos presuntivos.", styles['Normal']))
        _add_subsection_title("C. PLAN DE ESTUDIO")
        if planes_estudio: [story.append(Paragraph(f"<b>Acción:</b> {_format_val(plan.accion)} [<b>Estado:</b> {'Completado' if plan.completado else 'Pendiente'}]", styles['Normal'])) for plan in planes_estudio]
        else: story.append(Paragraph("No se registró un plan de estudio.", styles['Normal']))
        _add_subsection_title("ASESORAMIENTO Y EVOLUCIONES"); asesoramiento_text = "".join([f"<b>Fecha {_format_val(plan.fecha_visita)}:</b> {_format_val(plan.asesoramiento_evoluciones)}<br/><br/>" for plan in planes_estudio.filter(asesoramiento_evoluciones__isnull=False).exclude(asesoramiento_evoluciones__exact='')])
        story.append(Paragraph(asesoramiento_text if asesoramiento_text else "Sin evoluciones registradas.", styles['Normal']))

    # ===== INICIO DE LA MODIFICACIÓN 2: Sección de Autorización =====
    if autorizaciones:
        story.append(PageBreak())
        _add_section_title("AUTORIZACIÓN Y CONSENTIMIENTO INFORMADO")

        auth_intro = "El suscrito paciente o representante del mismo Sr. o Sra.:"
        auth_body = "Autorizo a él o los médicos del  Presente Instituto de Investigaciones Genéticas a efectuar todo examen físico y de laboratorio, fotografías y/o videos clínicos, imagenes diagnósticas, terapéuticas, estudios anatomopatológicos incluyendo necropsia en caso de muerte, y diversas técnicas de diagnositico prenatal que se consideren pertinentes para el diagnositico y tratamiento del mismo. Así mismo, autorizo a que la información, resultados de examenes complementarios, fotografías y/o videos obtenidos durante el estudio del paciente, siempre y cuando no se mencione su nombre, pueden ser divulgados por medios audiovisuales y/o publicaciones escritas con fines estrictamente académicas o científicos."
        
        for auto in autorizaciones:
            firmante_nombre = f"{auto.representante_padre.nombres} {auto.representante_padre.apellidos}" if auto.proposito.is_minor() and auto.representante_padre else f"{auto.proposito.nombres} {auto.proposito.apellidos}"
            firmante_id = f"C.I: {_format_val(auto.representante_padre.identificacion)}" if auto.proposito.is_minor() and auto.representante_padre else f"C.I: {_format_val(auto.proposito.identificacion)}"
            firmante_rol = "Representante" if auto.proposito.is_minor() and auto.representante_padre else "Paciente"
            firmante_display = f"{firmante_nombre} ({firmante_id}) - {firmante_rol}"
            
            story.append(Paragraph(f"<b>Consentimiento para:</b> {auto.proposito.nombres} {auto.proposito.apellidos}", styles['SubSectionTitle']))
            story.append(Spacer(1, 0.2 * inch))

            story.append(Paragraph(auth_intro, styles['Normal']))
            story.append(Spacer(1, 0.1 * inch))
            story.append(Paragraph(auth_body, styles['Justify']))

            story.append(Spacer(1, 0.4 * inch))
            
            if auto.archivo_autorizacion and hasattr(auto.archivo_autorizacion, 'path'):
                try:
                    signature_img = Image(auto.archivo_autorizacion.path, width=2.5 * inch, height=1.2 * inch)
                    signature_img.preserveAspectRatio = True
                    signature_img.hAlign = 'LEFT'
                    story.append(signature_img)
                except Exception as e:
                    print(f"Error al procesar imagen de firma para PDF (path: {auto.archivo_autorizacion.name}): {e}")
                    story.append(Paragraph("<i>[Firma no disponible - Error al procesar imagen]</i>", styles['Alert']))
                    story.append(Spacer(1, 0.4 * inch))
            else:
                story.append(Spacer(1, 1.2 * inch)) # Espacio para la firma manual

            story.append(Paragraph("______________________________", styles['Left']))
            story.append(Paragraph(f"<b>Firma de:</b> {firmante_display}", styles['Left']))
            story.append(Spacer(1, 0.5 * inch))
    # ===== FIN DE LA MODIFICACIÓN 2 =====

    try: doc.build(story)
    except Exception as e: print(f"CRITICAL PDF BUILD ERROR: {e}"); return HttpResponse(f"Error crítico al construir el PDF. Revise la consola del servidor. Error: {e}", status=500)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Historia_Clinica_{historia.numero_historia}.pdf"'
    return response

@require_POST
@login_required
@genetista_or_admin_required
def archive_historia(request, historia_id):
    """
    Archiva una historia clínica, guarda su estado previo y establece el motivo.
    """
    historia = get_object_or_404(HistoriasClinicas, pk=historia_id)
    user_profile = request.user.genetistas

    if user_profile.rol == 'GEN' and historia.genetista != user_profile:
        messages.error(request, "No tiene permiso para archivar esta historia clínica.")
        return redirect('ver_historias')

    form = ArchivarHistoriaForm(request.POST)
    if form.is_valid():
        motivo = form.cleaned_data['motivo']
        
        # --- LÓGICA MODIFICADA ---
        historia.estado_previo_archivado = historia.estado # Guardamos el estado actual
        historia.estado = HistoriasClinicas.ESTADO_ARCHIVADA
        historia.motivo_archivado = motivo
        historia.save() # El método save() del modelo se encarga de inactivar los pacientes
        messages.success(request, f"La historia clínica N° {historia.numero_historia} ha sido archivada.")
    else:
        messages.error(request, "Debe proporcionar un motivo para archivar la historia.")
    
    return redirect('ver_historias')

# ===== VISTA `unarchive_historia` MODIFICADA =====
@require_POST
@login_required
@admin_required
def unarchive_historia(request, historia_id):
    """
    Desarchiva una historia clínica, restaurando su estado previo.
    """
    historia = get_object_or_404(HistoriasClinicas, pk=historia_id)
    
    # --- LÓGICA MODIFICADA ---
    # Restauramos al estado guardado. Si no existe (p.ej. data antigua), usamos 'finalizada' como fallback.
    estado_restaurado = historia.estado_previo_archivado or HistoriasClinicas.ESTADO_FINALIZADA
    historia.estado = estado_restaurado
    historia.estado_previo_archivado = None # Limpiamos el estado previo
    historia.motivo_archivado = "" # Limpiamos el motivo
    historia.save() # El método save() del modelo se encarga de reactivar los pacientes
    
    messages.success(request, f"La historia clínica N° {historia.numero_historia} ha sido desarchivada y sus pacientes reactivados.")
    return redirect('ver_historias')

# ===== FUNCIÓN AUXILIAR `_get_pacientes_queryset_for_role` MODIFICADA =====
def _get_pacientes_queryset_for_role(user):
    """
    Función auxiliar para obtener el queryset de pacientes basado en el rol del usuario.
    ACTUALIZADO: Excluye pacientes inactivos para roles no-admin.
    """
    try:
        user_gen_profile = user.genetistas
    except Genetistas.DoesNotExist:
        if user.is_superuser:
            user_gen_profile, _ = Genetistas.objects.get_or_create(user=user, defaults={'rol': 'ADM'})
        else:
            return Propositos.objects.none()

    role = user_gen_profile.rol
    base_qs = Propositos.objects.filter(historia__isnull=False).select_related('historia', 'historia__genetista__user')

    # --- LÓGICA DE FILTRADO DE ROLES MODIFICADA ---
    # 1. Filtramos por Genetista si es necesario
    if role == 'GEN':
        base_qs = base_qs.filter(historia__genetista=user_gen_profile)
    elif role == 'LEC':
        if user_gen_profile.associated_genetista:
            base_qs = base_qs.filter(historia__genetista=user_gen_profile.associated_genetista)
        else:
            return Propositos.objects.none()
    
    # 2. Excluimos inactivos si el rol NO es ADM o Superuser
    if role != 'ADM' and not user.is_superuser:
        base_qs = base_qs.exclude(estado=Propositos.ESTADO_INACTIVO)
            
    return base_qs


@require_POST
@login_required
@genetista_or_admin_required
def guardar_diagnostico_final_ajax(request, evaluacion_id):
    """
    Vista AJAX para guardar o actualizar el diagnóstico final de una evaluación.
    """
    evaluacion = get_object_or_404(EvaluacionGenetica, pk=evaluacion_id)
    
    # Verificación de permisos (opcional pero recomendado)
    user_profile = request.user.genetistas
    if user_profile.rol == 'GEN':
        genetista_historia = None
        if evaluacion.proposito:
            genetista_historia = evaluacion.proposito.historia.genetista
        elif evaluacion.pareja:
            genetista_historia = evaluacion.pareja.proposito_id_1.historia.genetista
        if genetista_historia != user_profile:
            return JsonResponse({'success': False, 'message': 'No tiene permiso.'}, status=403)

    try:
        data = json.loads(request.body)
        diagnostico_texto = data.get('diagnostico_final', '').strip()

        evaluacion.diagnostico_final = diagnostico_texto
        evaluacion.save(update_fields=['diagnostico_final'])
        
        return JsonResponse({'success': True, 'message': 'Diagnóstico final guardado.'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Solicitud malformada.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error del servidor: {e}'}, status=500)



# --- Example/Tutorial Views ---
def hello(request, username): return JsonResponse({"message": f"Hello {username}"}) 
def about(request): return render(request, "about.html", {'username': request.user.username if request.user.is_authenticated else "Invitado"})
@login_required 
def projects(request): return render(request, "projects.html", {'projects': Project.objects.all()})
@login_required
def tasks(request): return render(request, "tasks.html", {'tasks': Task.objects.all()})
@login_required
def create_task(request):
    if request.method == 'POST':
        form = CreateNewTask(request.POST)
        if form.is_valid(): messages.success(request, "Tarea creada (ejemplo)."); return redirect('task_list') 
        else: messages.error(request, "Error en formulario de tarea (ejemplo).")
    else: form = CreateNewTask()
    return render(request, 'create_task.html', {'form': form})
@login_required
def create_project(request):
    if request.method == 'POST':
        form = CreateNewProject(request.POST)
        if form.is_valid(): messages.success(request, "Proyecto creado (ejemplo)."); return redirect('project_list') 
        else: messages.error(request, "Error en formulario de proyecto (ejemplo).")
    else: form = CreateNewProject()
    return render(request, 'create_project.html', {'form': form})
@login_required
def project_detail(request, id):
    project_instance = get_object_or_404(Project, id=id) 
    return render(request, 'detail.html', {'project': project_instance, 'tasks': Task.objects.filter(project=project_instance)})

@login_required
@all_roles_required
def flow_completion_view(request):
    """
    Vista simple que muestra una página de éxito al finalizar un flujo de formularios.
    """
    # Puedes añadir un mensaje genérico si lo deseas, aunque los mensajes
    # de las vistas anteriores ya deberían estar en la cola de mensajes.
    # messages.success(request, "Proceso completado exitosamente.")
    return render(request, 'flow_completion.html')


